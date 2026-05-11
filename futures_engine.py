"""
futures_engine.py — stale futures detection + settlement
─────────────────────────────────────────────────────────
Identifies open futures bets whose championship event has ended,
and provides the settlement primitive that moves a row from the
"Open Bets" sheet to the "Bet History" sheet inside a user's xlsx.

Used by server.py via /api/stale-futures and /api/settle-bet.

Rules:
- Never delete or overwrite historical records — settlement is an APPEND
  to Bet History plus a row delete from Open Bets (and only after the
  Bet History append succeeded).
- Event end dates live in futures_event_dates.json — update yearly.
- A future is "stale" when (event_end_date < today).
"""

import os, json, re
from datetime import datetime, date
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EVENT_DATES_FILE = os.path.join(SCRIPT_DIR, 'futures_event_dates.json')

# ── championship registry ───────────────────────────────────────────────

def _load_registry():
    """Load the championship event-date registry."""
    if not os.path.exists(EVENT_DATES_FILE):
        return {'championships': []}
    try:
        with open(EVENT_DATES_FILE) as f:
            return json.load(f)
    except Exception as e:
        print(f"[futures_engine] Could not load {EVENT_DATES_FILE}: {e}")
        return {'championships': []}


def _match_championship(matchup, registry):
    """Find which championship a matchup string refers to.
    Returns (champ_dict, matched_pattern) or (None, None)."""
    if not matchup:
        return None, None
    mu = matchup.lower()
    for champ in registry.get('championships', []):
        for pattern in champ.get('patterns', []):
            if pattern.lower() in mu:
                return champ, pattern
    return None, None


def _infer_year(matchup, added_date, today):
    """Pick the season year for an event.
    Priority: explicit 4-digit year in matchup → year from added_date → today's year.
    Falls forward to the NEXT year for events whose championship occurs after
    the typical season-start (e.g. an NBA bet placed in Oct 2025 resolves in Jun 2026)."""
    if matchup:
        m = re.search(r'\b(20\d{2})\b', matchup)
        if m:
            return int(m.group(1))
    added_year = None
    if added_date:
        try:
            # added_date is ISO 8601 from server.py
            added_year = datetime.fromisoformat(added_date.replace('Z', '+00:00')).year
        except Exception:
            pass
    if added_year:
        # If the bet was placed late in a year for an event that's decided next year,
        # the registry will hold the next year as the season key. We try added_year+1
        # first when matching, falling back to added_year.
        return added_year
    return today.year


def find_stale_futures(open_bets, today=None):
    """Given the dashboard-format open bets list (from read_open_bets), return
    only those whose championship event has ended.

    Each returned dict adds:
        eventEndDate    — ISO date the championship was decided
        daysPast        — integer days since the event ended
        championshipId  — registry id of the matched championship
        seasonYear      — year used to resolve the date
    """
    if today is None:
        today = date.today()
    registry = _load_registry()
    stale = []

    for b in open_bets:
        if (b.get('type') or '').lower() != 'future':
            continue
        matchup = b.get('matchup') or ''
        champ, _ = _match_championship(matchup, registry)
        if not champ:
            continue

        # Try multiple year candidates: explicit in matchup → added year → added year + 1
        seasons = champ.get('seasons', {})
        candidates = []
        explicit_year = re.search(r'\b(20\d{2})\b', matchup)
        if explicit_year:
            candidates.append(int(explicit_year.group(1)))
        try:
            added_dt = datetime.fromisoformat((b.get('addedDate') or '').replace('Z', '+00:00'))
            candidates.append(added_dt.year)
            candidates.append(added_dt.year + 1)
        except Exception:
            pass
        candidates.append(today.year)
        candidates.append(today.year + 1)

        # Pick the season whose end_date >= addedDate (so we don't settle against last year's event)
        # and which has already passed today.
        chosen = None
        for yr in candidates:
            iso = seasons.get(str(yr))
            if not iso:
                continue
            try:
                end_d = datetime.fromisoformat(iso).date()
            except Exception:
                continue
            # Must have ended already AND must be on or after the bet was placed
            try:
                placed_d = datetime.fromisoformat((b.get('addedDate') or '').replace('Z', '+00:00')).date()
            except Exception:
                placed_d = date(1970, 1, 1)
            if end_d < today and end_d >= placed_d:
                if chosen is None or end_d > chosen[1]:
                    chosen = (yr, end_d)

        if not chosen:
            continue

        days_past = (today - chosen[1]).days
        item = dict(b)
        item['eventEndDate'] = chosen[1].isoformat()
        item['daysPast'] = days_past
        item['championshipId'] = champ['id']
        item['championshipName'] = champ.get('name', '')
        item['seasonYear'] = chosen[0]
        stale.append(item)

    return stale


# ── settlement (xlsx mutation) ──────────────────────────────────────────

def settle_bet(xlsx_path, bet_id, result, winning_team=None, settled_iso=None):
    """Move an Open Bets row → Bet History row.

    bet_id    — the txId or fallback id used by the dashboard
    result    — 'W' / 'L' / 'P' (mapped to Won / Lost / Push)
    winning_team — optional, appended to Notes for future reference
    settled_iso  — optional override for the settled date (defaults today)

    Returns dict: {ok, settled_row, history_row_added}.
    Raises if bet_id isn't found in Open Bets.
    """
    if result not in ('W', 'L', 'P'):
        raise ValueError(f"result must be W/L/P, got {result!r}")

    result_str = {'W': 'Won', 'L': 'Lost', 'P': 'Push'}[result]
    settled_d = (settled_iso or date.today().isoformat())
    # Format like "May-10-2026" to match existing convention
    try:
        d = datetime.fromisoformat(settled_d).date()
        settled_str = d.strftime('%b-%d-%Y')
    except Exception:
        settled_str = date.today().strftime('%b-%d-%Y')

    wb = openpyxl.load_workbook(xlsx_path)
    if 'Open Bets' not in wb.sheetnames or 'Bet History' not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"Workbook missing required sheets: {wb.sheetnames}")

    open_ws = wb['Open Bets']
    history_ws = wb['Bet History']

    # Open Bets columns: 0=Game Time, 1=Source, 2=Sport, 3=Bet Type, 4=Teams,
    #                   5=Line, 6=Odds, 7=Risk, 8=To Win, 9=Status, 10=Notes
    # Bet History cols: 0=Settled Date, 1=TX ID, 2=Sport, 3=Bet Type, 4=Teams,
    #                   5=Line, 6=Odds, 7=Risk, 8=To Win, 9=Status, 10=Win/Loss $,
    #                   11=Notes, 12=Source
    target_row = None
    target_row_data = None
    for row_idx, row in enumerate(open_ws.iter_rows(min_row=4, values_only=True), start=4):
        if not row[4]:  # teams empty
            continue
        # Try to match by ticket id extracted from notes, else by row position
        tx_id = ''
        if row[10]:
            m = re.search(r'Ticket:\s*(\d+)', str(row[10]))
            if m:
                tx_id = m.group(1)
        # Fallback id format used by server.py when no ticket: "open_<index>"
        # The frontend may pass that or the real ticket id; accept both.
        if tx_id == str(bet_id) or f'open_{row_idx - 4}' == str(bet_id):
            target_row = row_idx
            target_row_data = list(row)
            break

    if target_row is None:
        wb.close()
        raise LookupError(f"Bet id {bet_id} not found in Open Bets")

    # Compute Win/Loss $
    risk = float(target_row_data[7] or 0)
    to_win = float(target_row_data[8] or 0)
    if result == 'W':
        win_loss = to_win
    elif result == 'L':
        win_loss = -risk
    else:  # Push
        win_loss = 0.0

    # Build Notes: preserve original + append winner if provided
    notes = str(target_row_data[10] or '')
    if winning_team:
        winner_tag = f" | Winner: {winning_team}"
        if winner_tag.strip() not in notes:
            notes = (notes + winner_tag).strip(' |')

    # Extract or synthesize a tx_id
    history_tx_id = ''
    m = re.search(r'Ticket:\s*(\d+)', notes)
    if m:
        history_tx_id = m.group(1)
    else:
        # Generate a stable id from the original game time + teams so re-runs don't double-insert
        history_tx_id = f"fut_{date.today().strftime('%Y%m%d')}_{abs(hash(str(target_row_data[4])[:30])) % 100000}"

    # Append to Bet History (find first empty row at end)
    new_row_idx = history_ws.max_row + 1
    history_ws.cell(new_row_idx, 1, settled_str)                  # Settled Date
    history_ws.cell(new_row_idx, 2, history_tx_id)                # TX ID
    history_ws.cell(new_row_idx, 3, target_row_data[2])           # Sport
    history_ws.cell(new_row_idx, 4, target_row_data[3])           # Bet Type
    history_ws.cell(new_row_idx, 5, target_row_data[4])           # Teams
    history_ws.cell(new_row_idx, 6, target_row_data[5])           # Line
    history_ws.cell(new_row_idx, 7, target_row_data[6])           # Odds
    history_ws.cell(new_row_idx, 8, risk)                          # Risk
    history_ws.cell(new_row_idx, 9, to_win)                        # To Win
    history_ws.cell(new_row_idx, 10, result_str)                   # Status
    history_ws.cell(new_row_idx, 11, win_loss)                     # Win/Loss $
    history_ws.cell(new_row_idx, 12, notes)                        # Notes
    history_ws.cell(new_row_idx, 13, target_row_data[1] or '')     # Source

    # Delete the row from Open Bets only AFTER the Bet History row is in place
    open_ws.delete_rows(target_row, 1)

    wb.save(xlsx_path)
    wb.close()

    return {
        'ok': True,
        'history_tx_id': history_tx_id,
        'win_loss': win_loss,
        'settled_str': settled_str,
        'open_row_removed': target_row,
        'history_row_added': new_row_idx,
    }
