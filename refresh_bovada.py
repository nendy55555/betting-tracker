"""
refresh_bovada.py
─────────────────
Logs into Bovada, pulls open bets and recent settled bets,
then updates Betting_Tracker.xlsx (Open Bets + Bet History sheets).

Open bets are merged with Locks25 open bets and sorted by game start time.

Usage:
    python3 refresh_bovada.py                # normal run
    python3 refresh_bovada.py --dry-run      # scrape but skip xlsx write
    python3 refresh_bovada.py --verbose      # DEBUG logging
    python3 refresh_bovada.py --no-preflight # skip Chrome preflight

Exit codes: 0=ok, 1=auth, 2=scrape, 3=browser, 4=excel. See scraper_common.py.
"""

import logging
import os
import re
import sys
import time
from datetime import datetime

from dotenv import load_dotenv
import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from scraper_common import (
    ScraperAuthError, ScraperBrowserError, ScraperExcelError,
    require_env, run_scraper, save_xlsx_safely, with_retry,
)

# ── credentials (loaded from .env; validated lazily in main) ────────────────
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))
BOVADA_URL = "https://www.bovada.lv"

# ── output file ──────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
TRACKER_PATH = os.path.join(SCRIPT_DIR, "Betting_Tracker.xlsx")

# ── style constants ───────────────────────────────────────────────────────────
DARK_BG    = "FF1A1A2E"
MID_BG     = "FF16213E"
GREEN_WIN  = "FF00C897"
RED_LOSS   = "FFFF5252"
LIGHT_ROW1 = "FFF8F9FA"
LIGHT_ROW2 = "FFECEFF4"
WHITE      = "FFFFFFFF"
BLACK      = "FF000000"

def side():
    return Side(style="thin", color="FFD0D0D0")

def bdr():
    s = side()
    return Border(left=s, right=s, top=s, bottom=s)

def hfill(color):
    return PatternFill("solid", fgColor=color)

def dfont(color=BLACK, size=9, bold=False):
    return Font(color=color, size=size, name="Arial", bold=bold)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ─────────────────────────────────────────────────────────────────────────────
# BROWSER
# ─────────────────────────────────────────────────────────────────────────────
def get_driver():
    opts = Options()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")
    try:
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()), options=opts)
    except Exception as e:
        raise ScraperBrowserError(f"Failed to start Chrome driver: {e}") from e
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def wait_for(driver, by, value, timeout=20):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value)))


# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────
@with_retry(retries=2)
def login(driver, email, password):
    logging.info("Navigating to Bovada...")
    driver.get(BOVADA_URL)
    time.sleep(3)

    # Click Login button
    try:
        login_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR,
                "[data-test='login-btn'], .login-btn, button[class*='login']")))
        login_btn.click()
        time.sleep(1)
    except Exception:
        # Try direct navigation as fallback
        driver.get(f"{BOVADA_URL}/en/sign-in")
        time.sleep(2)

    email_ok = pw_ok = submit_ok = False

    # Fill email/username
    try:
        email_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                "input[type='email'], input[name='email'], input[placeholder*='Email']")))
        email_field.clear()
        email_field.send_keys(email)
        email_ok = True
    except Exception as e:
        logging.warning("Email field not found: %s", e)

    # Fill password
    try:
        pw_field = driver.find_element(By.CSS_SELECTOR,
            "input[type='password'], input[name='password']")
        pw_field.clear()
        pw_field.send_keys(password)
        pw_ok = True
    except Exception as e:
        logging.warning("Password field not found: %s", e)

    # Submit
    try:
        submit_btn = driver.find_element(By.CSS_SELECTOR,
            "button[type='submit'], input[type='submit'], [data-test='submit-btn']")
        submit_btn.click()
        submit_ok = True
    except Exception as e:
        logging.warning("Submit btn not found, trying Enter key: %s", e)
        try:
            from selenium.webdriver.common.keys import Keys
            driver.find_element(By.CSS_SELECTOR, "input[type='password']").send_keys(Keys.RETURN)
            submit_ok = True
        except Exception:
            pass

    if not (email_ok and pw_ok and submit_ok):
        raise ScraperAuthError(
            "Bovada login form did not render as expected — "
            f"email={email_ok}, password={pw_ok}, submit={submit_ok}. "
            "Either creds are wrong or the site UI changed."
        )

    time.sleep(5)
    logging.info("Login attempted.")


# ─────────────────────────────────────────────────────────────────────────────
# SCRAPE OPEN BETS
# ─────────────────────────────────────────────────────────────────────────────
def scrape_open_bets(driver):
    """Navigate to Bovada My Bets / Active Bets section."""
    print("Fetching Bovada open bets...")
    bets = []
    open_bet_urls = [
        f"{BOVADA_URL}/mybets",
        f"{BOVADA_URL}/en/mybets",
        f"{BOVADA_URL}/account/mybets",
    ]
    for url in open_bet_urls:
        try:
            driver.get(url)
            time.sleep(4)
            page = driver.page_source
            if "bet" in page.lower() and len(page) > 5000:
                break
        except Exception:
            continue

    try:
        # Bovada uses Angular/React; look for bet cards
        bet_cards = driver.find_elements(By.CSS_SELECTOR,
            "[class*='bet-item'], [class*='pending-bet'], [class*='open-bet'], "
            "[data-test*='bet'], .bet-card, li[class*='bet']")

        for card in bet_cards:
            try:
                text = card.text.strip()
                if len(text) < 10:
                    continue

                lines = [l.strip() for l in text.split("\n") if l.strip()]
                description = " | ".join(lines[:4])
                risk_match  = re.search(r"\$[\d,]+\.?\d*", text)
                odds_match  = re.search(r"([+-]\d{3,4})", text)

                bets.append({
                    "date_placed":   lines[0] if lines else "",
                    "ticket":        f"BOV-{len(bets)+1}",
                    "description":   description,
                    "status":        "Pending",
                    "risk_win":      risk_match.group() if risk_match else "",
                    "odds":          odds_match.group() if odds_match else "",
                    "source":        "Bovada",
                    "game_start":    "",
                })
            except Exception:
                continue
    except Exception as e:
        print(f"  Bovada open bets error: {e}")

    print(f"  Found {len(bets)} Bovada open bets.")
    return bets


# ─────────────────────────────────────────────────────────────────────────────
# SCRAPE SETTLED BETS
# ─────────────────────────────────────────────────────────────────────────────
def scrape_settled_bets(driver):
    """Pull recent settled/graded bets from Bovada."""
    print("Fetching Bovada settled bets...")
    bets = []
    settled_urls = [
        f"{BOVADA_URL}/mybets?status=settled",
        f"{BOVADA_URL}/en/mybets/settled",
        f"{BOVADA_URL}/account/mybets?tab=settled",
    ]
    for url in settled_urls:
        try:
            driver.get(url)
            time.sleep(4)
            page = driver.page_source
            if ("won" in page.lower() or "lost" in page.lower()) and len(page) > 5000:
                break
        except Exception:
            continue

    try:
        bet_cards = driver.find_elements(By.CSS_SELECTOR,
            "[class*='bet-item'], [class*='settled-bet'], [class*='graded-bet'], "
            "[data-test*='bet'], .bet-card, li[class*='bet']")

        for card in bet_cards:
            try:
                text = card.text.strip()
                if len(text) < 10:
                    continue
                lower = text.lower()
                status = "Won" if "won" in lower else ("Lost" if "lost" in lower else "Settled")
                lines  = [l.strip() for l in text.split("\n") if l.strip()]

                risk_match = re.search(r"risk[:\s]*\$?([\d,]+\.?\d*)", text, re.IGNORECASE)
                win_match  = re.search(r"win[:\s]*\$?([\d,]+\.?\d*)", text, re.IGNORECASE)
                wl_match   = re.search(r"([+-]\$[\d,]+\.?\d*)", text)
                odds_match = re.search(r"([+-]\d{3,4})", text)

                bets.append({
                    "settled_date": lines[0] if lines else "",
                    "transaction":  f"BOV-S-{len(bets)+1}",
                    "description":  " | ".join(lines[:4]),
                    "status":       status,
                    "risk_win":     f"${risk_match.group(1) if risk_match else '0'}/{win_match.group(1) if win_match else '0'}",
                    "win_loss":     wl_match.group() if wl_match else "",
                    "odds":         odds_match.group() if odds_match else "",
                    "source":       "Bovada",
                })
            except Exception:
                continue
    except Exception as e:
        print(f"  Bovada settled bets error: {e}")

    print(f"  Found {len(bets)} Bovada settled bets.")
    return bets


# ─────────────────────────────────────────────────────────────────────────────
# GAME START TIMES (ESPN API)
# ─────────────────────────────────────────────────────────────────────────────
def fetch_game_times():
    print("Fetching game start times from ESPN...")
    game_times = {}
    sports = [
        ("basketball", "mens-college-basketball"),
        ("basketball", "nba"),
        ("soccer",     "usa.1"),
    ]
    headers = {"User-Agent": "Mozilla/5.0 Chrome/122.0.0.0"}
    for sport, league in sports:
        try:
            url = (f"https://site.api.espn.com/apis/site/v2/sports/"
                   f"{sport}/{league}/scoreboard")
            r = requests.get(url, headers=headers, timeout=10)
            data = r.json()
            for event in data.get("events", []):
                game_time_str = event.get("date", "")
                try:
                    game_dt = datetime.fromisoformat(game_time_str.replace("Z", "+00:00"))
                except Exception:
                    game_dt = None
                for comp in event.get("competitions", []):
                    for team in comp.get("competitors", []):
                        tname = team.get("team", {}).get("displayName", "").lower()
                        if tname and game_dt:
                            game_times[tname] = game_dt
        except Exception as e:
            print(f"  ESPN {league}: {e}")
    print(f"  Loaded {len(game_times)} schedules.")
    return game_times


def sort_by_game_time(bets, game_times):
    FAR = datetime(2099, 1, 1)

    def get_dt(bet):
        desc = bet.get("description", "").lower()
        for tname, dt in game_times.items():
            if tname in desc:
                bet["game_start"] = dt.strftime("%Y-%m-%d %H:%M UTC")
                return dt
        bet.setdefault("game_start", "Unknown")
        return FAR

    return sorted(bets, key=get_dt)


# ─────────────────────────────────────────────────────────────────────────────
# WRITE TO EXCEL — append Bovada bets (tagged with source)
# ─────────────────────────────────────────────────────────────────────────────
def merge_open_bets_into_sheet(wb, bovada_open, game_times):
    """Read existing Locks25 open bets, merge Bovada, re-sort, write back."""
    ws = wb["Open Bets"]

    # Collect existing Locks25 rows
    existing = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[1] or row[4]:  # source or description
            existing.append({
                "game_start":  str(row[0] or ""),
                "source":      str(row[1] or ""),
                "sport":       str(row[2] or ""),
                "bet_type":    str(row[3] or ""),
                "description": str(row[4] or ""),
                "line":        str(row[5] or ""),
                "odds":        str(row[6] or ""),
                "risk_win":    str(row[7] or ""),
                "to_win":      str(row[8] or ""),
                "status":      str(row[9] or "Pending"),
                "notes":       str(row[10] or ""),
            })

    all_bets = existing + [{
        "game_start":  b.get("game_start", ""),
        "source":      "Bovada",
        "sport":       "",
        "bet_type":    "Straight/Parlay",
        "description": b.get("description", ""),
        "line":        "",
        "odds":        b.get("odds", ""),
        "risk_win":    b.get("risk_win", ""),
        "to_win":      "",
        "status":      "Pending",
        "notes":       b.get("date_placed", ""),
    } for b in bovada_open]

    all_bets = sort_by_game_time(all_bets, game_times)

    # Clear rows 3+
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.fill  = PatternFill()

    # Write back
    for ri, b in enumerate(all_bets, 3):
        rw_fill = hfill(LIGHT_ROW1 if ri % 2 == 0 else LIGHT_ROW2)
        vals = [b["game_start"], b["source"], b["sport"], b["bet_type"],
                b["description"], b["line"], b["odds"], b["risk_win"],
                b["to_win"], b["status"], b["notes"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font  = dfont(size=9)
            c.fill  = rw_fill
            c.alignment = LEFT if ci == 5 else CENTER
            c.border = bdr()
        # Source color coding
        src_cell = ws.cell(row=ri, column=2)
        if b["source"] == "Locks25":
            src_cell.font = dfont(color="FF1A56DB", size=9, bold=True)
        else:
            src_cell.font = dfont(color="FF7C3AED", size=9, bold=True)

    ws["A1"].value = (
        f"OPEN BETS  —  Locks25 + Bovada  |  Sorted by Game Start  |  "
        f"Last refreshed: {datetime.now().strftime('%b-%d-%Y %H:%M')}"
    )
    print(f"  Open Bets sheet updated: {len(all_bets)} total open bets.")


def append_settled_to_history(wb, settled_bets):
    ws = wb["Bet History"]
    existing_txns = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[1]:
            existing_txns.add(str(row[1]))

    added = 0
    for b in settled_bets:
        txn = b.get("transaction", "")
        if txn in existing_txns:
            continue
        nr = ws.max_row + 1
        rw_fill = hfill(LIGHT_ROW1 if nr % 2 == 0 else LIGHT_ROW2)
        status = b.get("status", "")
        wl_raw = b.get("win_loss", "0").replace("$", "").replace(",", "").replace("+", "")
        try:
            wl_val = float(wl_raw)
        except Exception:
            wl_val = 0.0

        rw = b.get("risk_win","").split("/")
        try:
            risk = float(rw[0].replace("$","").replace(",",""))
        except Exception:
            risk = 0.0

        vals = [
            b.get("settled_date",""), txn,
            "CBB/NBA", "Straight",
            b.get("description",""), "", b.get("odds",""),
            risk, 0, status, wl_val, "Auto-imported from Bovada"
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=nr, column=ci, value=v)
            c.font  = dfont(size=9)
            c.fill  = rw_fill
            c.alignment = LEFT if ci in (5,12) else CENTER
            c.border = bdr()
        if status == "Won":
            ws.cell(row=nr, column=10).font = dfont(size=9, bold=True, color=GREEN_WIN)
        elif status == "Lost":
            ws.cell(row=nr, column=10).font = dfont(size=9, bold=True, color=RED_LOSS)
        ws.cell(row=nr, column=11).number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
        added += 1
        existing_txns.add(txn)

    print(f"  Added {added} new Bovada settled bets to Bet History.")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main(args):
    email, password = require_env("BOVADA_EMAIL", "BOVADA_PASSWORD")

    if not os.path.exists(TRACKER_PATH):
        raise ScraperExcelError(f"Tracker workbook not found: {TRACKER_PATH}")

    driver = get_driver()
    try:
        login(driver, email, password)
        bovada_open    = scrape_open_bets(driver)
        bovada_settled = scrape_settled_bets(driver)
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    game_times = fetch_game_times()
    bovada_open = sort_by_game_time(bovada_open, game_times)

    if args.dry_run:
        logging.info(
            "DRY-RUN — would merge %d open bets and append up to %d settled bets. "
            "Skipping xlsx write.", len(bovada_open), len(bovada_settled),
        )
        return

    wb = openpyxl.load_workbook(TRACKER_PATH)
    merge_open_bets_into_sheet(wb, bovada_open, game_times)
    append_settled_to_history(wb, bovada_settled)
    save_xlsx_safely(wb, TRACKER_PATH)
    logging.info("Bovada refresh complete. Tracker saved to: %s", TRACKER_PATH)


if __name__ == "__main__":
    sys.exit(run_scraper("bovada", main))
