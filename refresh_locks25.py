"""
refresh_locks25.py
──────────────────
Logs into Locks25, pulls open bets and this week's settled bets,
then updates Betting_Tracker.xlsx (Open Bets + Bet History sheets).

Open bets are sorted chronologically by game start time pulled from ESPN.

Usage:
    python3 refresh_locks25.py                # normal run
    python3 refresh_locks25.py --dry-run      # scrape but skip xlsx write
    python3 refresh_locks25.py --verbose      # DEBUG logging
    python3 refresh_locks25.py --no-preflight # skip Chrome preflight

Exit codes: 0=ok, 1=auth, 2=scrape, 3=browser, 4=excel. See scraper_common.py.
"""

import logging
import os
import re
import sys
import time
from datetime import datetime

from dotenv import load_dotenv
import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from scraper_common import (
    ScraperAuthError, ScraperBrowserError, ScraperDOMError, ScraperError,
    require_env, run_scraper, save_xlsx_safely, with_retry,
)

# ── credentials (loaded from .env; validated lazily in main) ────────────────
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))
LOCKS25_URL = "https://locks25.com"

# ── output file ──────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TRACKER_PATH = os.path.join(SCRIPT_DIR, "Betting_Tracker.xlsx")

# ── style constants ───────────────────────────────────────────────────────────
DARK_BG    = "FF1A1A2E"
MID_BG     = "FF16213E"
GREEN_WIN  = "FF00C897"
RED_LOSS   = "FFFF5252"
YELLOW_P   = "FFFFC107"
LIGHT_ROW1 = "FFF8F9FA"
LIGHT_ROW2 = "FFECEFF4"
WHITE      = "FFFFFFFF"
BLACK      = "FF000000"
ACCENT     = "FFE94560"

def side():
    return Side(style="thin", color="FFD0D0D0")

def bdr():
    s = side()
    return Border(left=s, right=s, top=s, bottom=s)

def hfill(color):
    return PatternFill("solid", fgColor=color)

def hfont(bold=True, color="FFFFFFFF", size=9):
    return Font(bold=bold, color=color, size=size, name="Arial")

def dfont(color=BLACK, size=9, bold=False):
    return Font(color=color, size=size, name="Arial", bold=bold)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ─────────────────────────────────────────────────────────────────────────────
# BROWSER SETUP
# ─────────────────────────────────────────────────────────────────────────────
def get_driver():
    opts = Options()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")
    try:
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()), options=opts)
    except Exception as e:
        raise ScraperBrowserError(f"Failed to start Chrome driver: {e}") from e
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def wait_for(driver, by, value, timeout=15):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value)))


# ─────────────────────────────────────────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────────────────────────────────────────
@with_retry(retries=2)
def login(driver, username, password):
    logging.info("Navigating to Locks25...")
    driver.get(LOCKS25_URL)
    time.sleep(2)

    try:
        wait_for(driver, By.CSS_SELECTOR, "input[placeholder='Username']").send_keys(username)
        driver.find_element(By.CSS_SELECTOR, "input[placeholder='Password']").send_keys(password)
        driver.find_element(By.XPATH, "//button[contains(text(),'LOGIN')]").click()
    except Exception as e:
        raise ScraperAuthError(
            f"Locks25 login form did not render as expected — likely DOM change or site down: {e}"
        ) from e
    time.sleep(3)

    # Post-login sanity: if we're still on a page with a visible LOGIN button, creds were rejected
    try:
        still_login = driver.find_elements(By.XPATH, "//button[contains(text(),'LOGIN')]")
        if still_login and any(b.is_displayed() for b in still_login):
            raise ScraperAuthError(
                "Locks25 login appears rejected — LOGIN button still visible after submit. "
                "Check LOCKS25_USERNAME / LOCKS25_PASSWORD in .env."
            )
    except ScraperAuthError:
        raise
    except Exception:
        pass  # DOM lookup failure is not evidence of auth failure

    logging.info("Logged in to Locks25.")


# ─────────────────────────────────────────────────────────────────────────────
# SCRAPE OPEN BETS
# ─────────────────────────────────────────────────────────────────────────────
def scrape_open_bets(driver):
    """Navigate to Open Bets page and extract all pending wagers."""
    print("Fetching open bets...")
    driver.get(f"{LOCKS25_URL}/Betting/OpenBets")
    time.sleep(3)

    bets = []
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "tr.bet-row, [class*='bet-item'], tr")
        for row in rows:
            try:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) >= 5:
                    date_placed = cells[0].text.strip()
                    ticket      = cells[1].text.strip()
                    description = cells[2].text.strip()
                    status      = cells[3].text.strip()
                    risk_win    = cells[4].text.strip() if len(cells) > 4 else ""

                    if ticket and re.match(r"6\d{8}", ticket):
                        bets.append({
                            "date_placed": date_placed,
                            "ticket": ticket,
                            "description": description,
                            "status": status,
                            "risk_win": risk_win,
                            "source": "Locks25",
                        })
            except Exception:
                continue
    except Exception as e:
        print(f"  Open bets scrape error: {e}")

    print(f"  Found {len(bets)} open bets.")
    return bets


# ─────────────────────────────────────────────────────────────────────────────
# SCRAPE SETTLED BETS (THIS WEEK + LAST WEEK)
# ─────────────────────────────────────────────────────────────────────────────
def _scrape_week(driver, week_value, label):
    """Scrape settled bets for a single week selection."""
    try:
        sel = Select(driver.find_element(
            By.CSS_SELECTOR, "select[class*='week'], select"))
        sel.select_by_value(str(week_value))
        time.sleep(2)
    except Exception:
        pass

    bets = []
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "tr, [class*='history-row']")
        for row in rows:
            try:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) >= 6:
                    settled     = cells[0].text.strip()
                    transaction = cells[1].text.strip()
                    description = cells[2].text.strip()
                    status      = cells[4].text.strip()
                    risk_win    = cells[5].text.strip() if len(cells) > 5 else ""
                    win_loss    = cells[6].text.strip() if len(cells) > 6 else ""

                    if transaction and re.match(r"6\d{8}", transaction):
                        bets.append({
                            "settled_date": settled,
                            "transaction":  transaction,
                            "description":  description,
                            "status":       status,
                            "risk_win":     risk_win,
                            "win_loss":     win_loss,
                            "source":       "Locks25",
                        })
            except Exception:
                continue
    except Exception as e:
        print(f"  History scrape error ({label}): {e}")

    print(f"  Found {len(bets)} settled bets ({label}).")
    return bets


def scrape_history_this_week(driver):
    """Navigate to History and pull this week + last week data."""
    print("Fetching settled bet history...")
    driver.get(f"{LOCKS25_URL}/Betting/History")
    time.sleep(3)

    bets = _scrape_week(driver, 0, "this week")
    bets += _scrape_week(driver, 1, "last week")

    seen = set()
    deduped = []
    for b in bets:
        if b["transaction"] not in seen:
            seen.add(b["transaction"])
            deduped.append(b)

    print(f"  Total unique settled bets: {len(deduped)}")
    return deduped


# ─────────────────────────────────────────────────────────────────────────────
# GAME START TIMES  (ESPN)
# ─────────────────────────────────────────────────────────────────────────────
def fetch_game_times():
    """
    Pulls today's game times from ESPN for CBB, NBA, NFL, Soccer.
    Returns dict: team_name (lower) -> ISO datetime string.
    """
    print("Fetching game start times from ESPN...")
    game_times = {}
    sports = [
        ("basketball", "mens-college-basketball"),
        ("basketball", "nba"),
        ("soccer",     "usa.1"),
    ]
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36"
    }
    for sport, league in sports:
        try:
            url = f"https://site.api.espn.com/apis/site/v2/sports/{sport}/{league}/scoreboard"
            r = requests.get(url, headers=headers, timeout=10)
            data = r.json()
            for event in data.get("events", []):
                game_time_str = event.get("date", "")
                try:
                    game_dt = datetime.fromisoformat(game_time_str.replace("Z", "+00:00"))
                except Exception:
                    game_dt = None
                for comp in event.get("competitions", []):
                    for team in comp.get("competitors", []):
                        tname = team.get("team", {}).get("displayName", "").lower()
                        if tname and game_dt:
                            game_times[tname] = game_dt
        except Exception as e:
            print(f"  ESPN error ({league}): {e}")

    print(f"  Loaded {len(game_times)} team schedules.")
    return game_times


def extract_teams(description):
    """Try to pull team names from a bet description."""
    # Pattern: "TeamA vs TeamB" or "[CBB] - TeamName ..."
    m = re.search(r"Basketball (.+?) vs (.+?) -", description)
    if m:
        return m.group(1).strip().lower(), m.group(2).strip().lower()
    # CBB format: "TEAM NAME -spread"
    m2 = re.search(r"\] (.+?) [+-][\d½]", description)
    if m2:
        return m2.group(1).strip().lower(), None
    return None, None


def parse_matchup(description):
    """Extract a clean 'Team A vs Team B' matchup string from a Locks25 bet description.
    Falls back to a single team name if only one can be found.
    Used to populate the Teams/Event column in Excel so the dashboard can display opponents.

    Locks25 description formats seen:
      "College Basketball Duke Blue Devils vs St. John's Red Storm - Handicap: +3 (-110)"
      "College Basketball [CBB] Arkansas Razorbacks +8 (-110)"
      "NBA Oklahoma City Thunder vs Denver Nuggets - Money Line"
      "STRAIGHT BET - Arkansas ATS +8"
    """
    d = description.strip()

    # Format 1: "Sport Team A vs Team B - ..."
    m = re.search(
        r'(?:College\s+)?(?:Basketball|Football|Baseball|Hockey|Soccer)\s+'
        r'(.+?)\s+vs\.?\s+(.+?)\s*[-–]',
        d, re.IGNORECASE
    )
    if m:
        return f"{m.group(1).strip()} vs {m.group(2).strip()}"

    # Format 2: bare "Team A vs Team B" anywhere in the string
    m2 = re.search(r'(.+?)\s+vs\.?\s+(.+?)(?:\s*[-–(]|$)', d, re.IGNORECASE)
    if m2:
        team_a = m2.group(1).strip()
        team_b = m2.group(2).strip()
        # Sanity check: ignore if either side looks like an odds value
        if not re.match(r'^[+-]?\d+$', team_a) and not re.match(r'^[+-]?\d+$', team_b):
            return f"{team_a} vs {team_b}"

    # Format 3: "[CBB] Team Name +spread" — only one team available
    m3 = re.search(r'\]\s*(.+?)\s+[+-][\d½¼¾]', d)
    if m3:
        return m3.group(1).strip()

    # Format 4: strip sport prefix and trailing odds/line, return what's left
    cleaned = re.sub(
        r'^(?:College\s+)?(?:Basketball|Football|Baseball|Hockey|Soccer|NBA|NFL|CBB)\s*',
        '', d, flags=re.IGNORECASE
    )
    cleaned = re.sub(r'\s+[+-][\d½¼¾.]+.*$', '', cleaned).strip()
    if cleaned and len(cleaned) > 2:
        return cleaned

    # Last resort: return the full description as-is
    return d


def sort_open_bets_by_time(open_bets, game_times):
    """Attach game start times and sort chronologically."""
    FAR_FUTURE = datetime(2099, 1, 1)

    def get_time(bet):
        desc = bet.get("description", "")
        t1, t2 = extract_teams(desc)
        for t in [t1, t2]:
            if t:
                for key, dt in game_times.items():
                    if t in key or key in t:
                        bet["game_start"] = dt.strftime("%Y-%m-%d %H:%M UTC")
                        return dt
        # Fall back: parse date_placed as rough proxy
        bet["game_start"] = "Unknown"
        return FAR_FUTURE

    open_bets.sort(key=get_time)
    return open_bets


# ─────────────────────────────────────────────────────────────────────────────
# WRITE TO EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def update_open_bets_sheet(wb, open_bets):
    ws = wb["Open Bets"]
    # Clear existing data (keep header rows 1-2)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.fill  = PatternFill()

    if not open_bets:
        ws.cell(row=3, column=1,
                value="No open bets found.").font = Font(italic=True, size=9, name="Arial")
        return

    for ri, b in enumerate(open_bets, 3):
        rw_fill = hfill(LIGHT_ROW1 if ri % 2 == 0 else LIGHT_ROW2)
        vals = [
            b.get("game_start", ""),
            b.get("source", ""),
            "CBB/NBA",
            "Straight/Parlay",
            parse_matchup(b.get("description", "")),   # "Team A vs Team B" format
            "",
            "",
            b.get("risk_win", ""),
            "",
            b.get("status", "Pending"),
            b.get("date_placed", ""),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font  = dfont(size=9)
            c.fill  = rw_fill
            c.alignment = LEFT if ci == 5 else CENTER
            c.border = bdr()

    ws["A1"].value = (
        f"OPEN BETS  —  Sorted by Game Start Time  |  "
        f"Last refreshed: {datetime.now().strftime('%b-%d-%Y %H:%M')}"
    )


def update_history_sheet(wb, settled_bets):
    """Append new settled bets to Bet History sheet (skip existing transaction IDs)."""
    ws = wb["Bet History"]

    # Collect existing transaction IDs
    existing = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[1]:
            existing.add(str(row[1]))

    added = 0
    for b in settled_bets:
        txn = b.get("transaction", "")
        if txn in existing:
            continue
        nr = ws.max_row + 1
        rw_fill = hfill(LIGHT_ROW1 if nr % 2 == 0 else LIGHT_ROW2)
        status = b.get("status", "")
        wl_raw = b.get("win_loss", "0").replace("$","").replace(",","").replace("+","")
        try:
            wl_val = float(wl_raw)
        except Exception:
            wl_val = 0.0

        rw_raw = b.get("risk_win","").split("/")
        try:
            risk = float(rw_raw[0].replace("$","").replace(",",""))
        except Exception:
            risk = 0.0
        try:
            to_win = float(rw_raw[1].replace("$","").replace(",",""))
        except Exception:
            to_win = 0.0

        vals = [
            b.get("settled_date",""), txn,
            "CBB/NBA/SOC", "Straight",
            parse_matchup(b.get("description","")), "", "",   # "Team A vs Team B" format
            risk, to_win, status, wl_val, "Auto-imported from Locks25",
            "Locks"
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=nr, column=ci, value=v)
            c.font  = dfont(size=9)
            c.fill  = rw_fill
            c.alignment = LEFT if ci in (5, 12) else CENTER
            c.border = bdr()
        if status == "Won":
            ws.cell(row=nr, column=10).font = dfont(size=9, bold=True, color=GREEN_WIN)
        elif status == "Lost":
            ws.cell(row=nr, column=10).font = dfont(size=9, bold=True, color=RED_LOSS)
        ws.cell(row=nr, column=11).number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
        added += 1

    print(f"  Added {added} new settled bets to Bet History.")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main(args):
    username, password = require_env("LOCKS25_USERNAME", "LOCKS25_PASSWORD")

    if not os.path.exists(TRACKER_PATH):
        # Workbook write happens at end; if the file is missing there's nothing to update.
        from scraper_common import ScraperExcelError
        raise ScraperExcelError(f"Tracker workbook not found: {TRACKER_PATH}")

    driver = get_driver()
    try:
        login(driver, username, password)
        open_bets    = scrape_open_bets(driver)
        settled_bets = scrape_history_this_week(driver)
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    game_times = fetch_game_times()
    open_bets  = sort_open_bets_by_time(open_bets, game_times)

    if args.dry_run:
        logging.info(
            "DRY-RUN — would update %d open bets and append up to %d settled bets. "
            "Skipping xlsx write.", len(open_bets), len(settled_bets),
        )
        return

    wb = openpyxl.load_workbook(TRACKER_PATH)
    update_open_bets_sheet(wb, open_bets)
    update_history_sheet(wb, settled_bets)
    save_xlsx_safely(wb, TRACKER_PATH)
    logging.info("Locks25 refresh complete. Tracker saved to: %s", TRACKER_PATH)


if __name__ == "__main__":
    sys.exit(run_scraper("locks25", main))
