"""
parse_bovada_paste.py
─────────────────────
Parses raw copy-pasted text from Bovada's Settled Bets page
and imports the bets into Betting_Tracker.xlsx.

Usage:
    python parse_bovada_paste.py <paste_file.txt>

    Or import and call: parse_bovada_text(raw_text) -> list of bet dicts
"""

import re
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── output file ──────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
TRACKER_PATH = os.path.join(SCRIPT_DIR, "Betting_Tracker.xlsx")

# ── style constants (match existing tracker) ─────────────────────────────────
GREEN_WIN  = "FF00C897"
RED_LOSS   = "FFFF5252"
YELLOW_PUSH = "FFFFC107"
LIGHT_ROW1 = "FFF8F9FA"
LIGHT_ROW2 = "FFECEFF4"
WHITE      = "FFFFFFFF"
BLACK      = "FF000000"

def side():
    return Side(style="thin", color="FFD0D0D0")

def bdr():
    s = side()
    return Border(left=s, right=s, top=s, bottom=s)

def hfill(color):
    return PatternFill("solid", fgColor=color)

def dfont(color=BLACK, size=9, bold=False):
    return Font(color=color, size=size, name="Arial", bold=bold)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─────────────────────────────────────────────────────────────────────────────
# PARSER
# ─────────────────────────────────────────────────────────────────────────────

def clean_bovada_text(raw_text):
    """Strip the header/footer junk from a Bovada page paste."""
    # Find where settled bets start (first date pattern after SETTLED or OPEN BETS)
    lines = raw_text.split("\n")
    cleaned = []
    in_bets = False
    for line in lines:
        stripped = line.strip()
        # Detect start of bet data: a date line like "3/23/26 12:18 AM"
        if re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}\s+\d{1,2}:\d{2}\s*(AM|PM)$", stripped, re.IGNORECASE):
            in_bets = True
        # Stop at footer markers
        if stripped.startswith("BET SLIP") or stripped.startswith("Playable Balance"):
            break
        if in_bets:
            cleaned.append(stripped)
    return "\n".join(cleaned)


def split_into_bet_blocks(text):
    """Split cleaned text into individual bet blocks.
    Each bet starts with a date line followed by a Ref. line.
    Game dates within a bet (not followed by Ref.) stay in the same block."""
    lines = text.split("\n")
    blocks = []
    current = []
    date_pattern = re.compile(r"^\d{1,2}/\d{1,2}/\d{2,4}\s+\d{1,2}:\d{2}\s*(AM|PM)$", re.IGNORECASE)

    for idx, line in enumerate(lines):
        stripped = line.strip()
        if not stripped:
            continue

        # A date line is a new bet ONLY if the next non-empty line starts with "Ref."
        if date_pattern.match(stripped):
            # Look ahead for Ref. line
            next_line = ""
            for j in range(idx + 1, len(lines)):
                if lines[j].strip():
                    next_line = lines[j].strip()
                    break
            if next_line.startswith("Ref."):
                # This is a new bet block
                if current:
                    blocks.append(current)
                current = [stripped]
                continue

        current.append(stripped)

    if current:
        blocks.append(current)

    return blocks


def detect_sport(event_text):
    """Guess the sport from event/team names."""
    lower = event_text.lower()

    # Soccer indicators
    soccer_teams = ["everton", "chelsea", "arsenal", "liverpool", "manchester",
                    "barcelona", "real madrid", "tottenham", "leicester"]
    for team in soccer_teams:
        if team in lower:
            return "Soccer"
    if "fc" in lower or "vs " in lower.split("@")[0] if "@" not in lower else False:
        pass

    # NCAA tournament seeds like (#5) suggest college basketball
    if re.search(r"\(#\d+\)", event_text):
        return "CBB"

    # NBA teams
    nba_teams = ["lakers", "celtics", "warriors", "nets", "knicks", "heat",
                 "bucks", "76ers", "suns", "nuggets", "clippers", "rockets",
                 "mavericks", "grizzlies", "cavaliers", "thunder", "pelicans",
                 "hawks", "bulls", "pistons", "pacers", "magic", "raptors",
                 "hornets", "wizards", "blazers", "kings", "spurs", "timberwolves",
                 "jazz"]
    for team in nba_teams:
        if team in lower:
            return "NBA"

    # Default
    return "CBB"


def parse_odds_value(odds_str):
    """Convert American odds string to numeric."""
    odds_str = odds_str.strip().replace(",", "")
    try:
        return int(odds_str)
    except ValueError:
        return 0


def calculate_to_win(risk, odds):
    """Calculate potential win from risk amount and American odds."""
    if odds > 0:
        return round(risk * (odds / 100), 2)
    elif odds < 0:
        return round(risk * (100 / abs(odds)), 2)
    return 0.0


def parse_bet_block(lines):
    """Parse a single bet block into a structured dict."""
    if len(lines) < 4:
        return None

    bet = {}

    # Line 0: settled date "3/23/26 12:18 AM"
    settled_date_raw = lines[0]
    try:
        dt = datetime.strptime(settled_date_raw, "%m/%d/%y %I:%M %p")
        bet["settled_date"] = dt.strftime("%b-%d-%Y")
        bet["settled_datetime"] = dt
    except ValueError:
        bet["settled_date"] = settled_date_raw
        bet["settled_datetime"] = None

    # Line 1: "Ref.26036669237570"
    ref_line = lines[1] if len(lines) > 1 else ""
    ref_match = re.match(r"Ref\.(\d+)", ref_line)
    bet["ref_id"] = ref_match.group(1) if ref_match else ref_line

    # Line 2: bet type "Single" or "2 Team Parlay" etc.
    bet_type_line = lines[2] if len(lines) > 2 else ""
    bet["bet_type_raw"] = bet_type_line

    if "parlay" in bet_type_line.lower():
        bet["is_parlay"] = True
        parlay_match = re.match(r"(\d+)\s+Team\s+Parlay", bet_type_line, re.IGNORECASE)
        bet["parlay_legs"] = int(parlay_match.group(1)) if parlay_match else 2
        bet["bet_type"] = f"{bet['parlay_legs']}-Leg Parlay"
    else:
        bet["is_parlay"] = False
        bet["parlay_legs"] = 1
        bet["bet_type"] = "Straight"

    # Line 3: status "WIN" or "LOSS" or "PUSH"
    status_line = lines[3] if len(lines) > 3 else ""
    status_upper = status_line.upper().strip()
    if status_upper == "WIN":
        bet["status"] = "Won"
    elif status_upper == "LOSS":
        bet["status"] = "Lost"
    elif status_upper == "PUSH":
        bet["status"] = "Push"
    elif status_upper == "CASHED OUT":
        bet["status"] = "Cashed Out"
    else:
        bet["status"] = status_upper

    # Remaining lines: parse legs, RISK, ODDS, WINNINGS
    remaining = lines[4:]
    bet["legs"] = []
    bet["risk"] = 0.0
    bet["odds"] = ""
    bet["winnings"] = 0.0

    i = 0
    current_leg = {}
    while i < len(remaining):
        line = remaining[i]

        # Event line: "* Texas Tech (#5) @ Alabama (#4)"
        if line.startswith("*") and ("@" in line or "vs" in line.lower()):
            # Save previous leg if exists
            if current_leg.get("event"):
                bet["legs"].append(current_leg)
            event = line.lstrip("* ").strip()
            current_leg = {"event": event, "pick": "", "game_date": ""}
            i += 1
            continue

        # Game date line: "3/22/26 9:45 PM"
        if re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}\s+\d{1,2}:\d{2}\s*(AM|PM)$", line, re.IGNORECASE):
            if current_leg:
                current_leg["game_date"] = line
            i += 1
            continue

        # Pick/selection line (comes after event or game date)
        # e.g. "Over 164.5 (-110) (Game) Total"
        # e.g. "Texas Tech (#12) +12.0 (-115) (Live Game) Point Spread"
        # e.g. "Iowa State (#2) (-235) (Game) Moneyline"
        if current_leg.get("event") and not current_leg.get("pick"):
            if not line.startswith("RISK") and not line.startswith("ODDS") and not line.startswith("WINNINGS"):
                if not line.startswith("$") and not line.startswith("+"):
                    # This looks like a pick line
                    current_leg["pick"] = line
                    i += 1
                    continue

        # RISK line
        if line == "RISK":
            i += 1
            if i < len(remaining):
                risk_match = re.search(r"\$\s*([\d,]+\.?\d*)", remaining[i])
                if risk_match:
                    bet["risk"] = float(risk_match.group(1).replace(",", ""))
            i += 1
            continue

        # ODDS line
        if line == "ODDS":
            i += 1
            if i < len(remaining):
                bet["odds"] = remaining[i].strip()
            i += 1
            continue

        # WINNINGS or CASHED OUT line
        if line in ("WINNINGS", "CASHED OUT"):
            i += 1
            if i < len(remaining):
                win_match = re.search(r"\$\s*([\d,]+\.?\d*)", remaining[i])
                if win_match:
                    bet["winnings"] = float(win_match.group(1).replace(",", ""))
            i += 1
            continue

        i += 1

    # Save last leg
    if current_leg.get("event"):
        bet["legs"].append(current_leg)

    # Build description and line/spread from legs
    if bet["legs"]:
        if bet["is_parlay"]:
            leg_descs = []
            for leg in bet["legs"]:
                pick = leg.get("pick", "")
                event = leg.get("event", "")
                leg_descs.append(f"{event}: {pick}" if pick else event)
            bet["description"] = " / ".join(leg_descs)
            bet["line_spread"] = f"{len(bet['legs'])}-leg parlay"
        else:
            leg = bet["legs"][0]
            bet["description"] = leg.get("event", "")
            bet["line_spread"] = leg.get("pick", "")
    else:
        bet["description"] = " ".join(remaining[:3]) if remaining else ""
        bet["line_spread"] = ""

    # Detect sport from first leg
    first_event = bet["legs"][0]["event"] if bet["legs"] else bet["description"]
    bet["sport"] = detect_sport(first_event)

    # Calculate win/loss amount
    if bet["status"] == "Won":
        bet["win_loss"] = round(bet["winnings"] - bet["risk"], 2)
    elif bet["status"] == "Lost":
        bet["win_loss"] = -bet["risk"]
    elif bet["status"] == "Cashed Out":
        bet["win_loss"] = round(bet["winnings"] - bet["risk"], 2)
    else:
        bet["win_loss"] = 0.0

    # Calculate to_win
    odds_val = parse_odds_value(bet["odds"])
    if bet["winnings"] > 0:
        bet["to_win"] = round(bet["winnings"] - bet["risk"], 2)
    else:
        bet["to_win"] = calculate_to_win(bet["risk"], odds_val)

    return bet


def parse_bovada_text(raw_text):
    """Main entry: parse raw Bovada paste into list of bet dicts."""
    cleaned = clean_bovada_text(raw_text)
    blocks = split_into_bet_blocks(cleaned)
    bets = []
    for block in blocks:
        parsed = parse_bet_block(block)
        if parsed:
            bets.append(parsed)
    return bets


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────────────────────────────────────

def get_existing_ref_ids(ws):
    """Collect all transaction IDs already in Bet History to prevent dupes."""
    existing = set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[1]:
            existing.add(str(row[1]))
    return existing


def write_bets_to_tracker(bets, tracker_path=None):
    """Append parsed Bovada bets to the Bet History sheet."""
    path = tracker_path or TRACKER_PATH
    wb = openpyxl.load_workbook(path)
    ws = wb["Bet History"]

    existing_refs = get_existing_ref_ids(ws)

    # Find TOTALS row if it exists, insert before it
    totals_row = None
    for r in range(ws.max_row, 2, -1):
        val = ws.cell(r, 1).value
        if val and str(val).upper().startswith("TOTAL"):
            totals_row = r
            break

    added = 0
    skipped = 0
    parlay_legs_added = 0

    for bet in bets:
        ref = bet["ref_id"]
        if ref in existing_refs:
            skipped += 1
            continue

        # Determine insert row
        if totals_row:
            ws.insert_rows(totals_row)
            nr = totals_row
            totals_row += 1
        else:
            nr = ws.max_row + 1

        rw_fill = hfill(LIGHT_ROW1 if nr % 2 == 0 else LIGHT_ROW2)

        # Build notes
        notes_parts = ["Bovada import"]
        if bet.get("is_parlay") and bet.get("legs"):
            leg_details = []
            for idx, leg in enumerate(bet["legs"], 1):
                leg_details.append(f"Leg {idx}: {leg.get('event', '')} - {leg.get('pick', '')}")
            notes_parts.append(" | ".join(leg_details))
        notes = "; ".join(notes_parts)

        vals = [
            bet["settled_date"],           # Settled Date
            ref,                           # Transaction ID
            bet["sport"],                  # Sport
            bet["bet_type"],               # Bet Type
            bet["description"],            # Teams / Event
            bet["line_spread"],            # Line / Spread
            bet["odds"],                   # Odds
            bet["risk"],                   # Risk ($)
            bet["to_win"],                 # To Win ($)
            bet["status"],                 # Status
            bet["win_loss"],               # Win/Loss ($)
            notes,                         # Notes
            "Bovada",                      # Source
        ]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=nr, column=ci, value=v)
            c.font = dfont(size=9)
            c.fill = rw_fill
            c.alignment = LEFT if ci in (5, 6, 12) else CENTER
            c.border = bdr()

        # Dollar formatting
        for col in [8, 9]:
            ws.cell(row=nr, column=col).number_format = '"$"#,##0.00'
        ws.cell(row=nr, column=11).number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'

        # Status color
        status_cell = ws.cell(row=nr, column=10)
        if bet["status"] == "Won":
            status_cell.font = dfont(size=9, bold=True, color=GREEN_WIN)
        elif bet["status"] == "Lost":
            status_cell.font = dfont(size=9, bold=True, color=RED_LOSS)
        elif bet["status"] in ("Push", "Cashed Out"):
            status_cell.font = dfont(size=9, bold=True, color=YELLOW_PUSH)

        added += 1
        existing_refs.add(ref)

        # Also add parlay legs to Parlays sheet if it's a parlay
        if bet["is_parlay"] and "Parlays" in wb.sheetnames:
            pw = wb["Parlays"]
            pnr = pw.max_row + 1
            p_fill = hfill(LIGHT_ROW1 if pnr % 2 == 0 else LIGHT_ROW2)

            leg_text = []
            for idx, leg in enumerate(bet["legs"], 1):
                leg_text.append(f"Leg {idx}: {leg.get('event', '')} - {leg.get('pick', '')}")

            p_vals = [
                bet["settled_date"],
                ref,
                bet["description"][:80],
                bet["risk"],
                bet["to_win"],
                bet["status"],
                bet["win_loss"],
                " | ".join(leg_text),
            ]
            for ci, v in enumerate(p_vals, 1):
                c = pw.cell(row=pnr, column=ci, value=v)
                c.font = dfont(size=9)
                c.fill = p_fill
                c.alignment = LEFT if ci in (3, 8) else CENTER
                c.border = bdr()
            pw.cell(row=pnr, column=4).number_format = '"$"#,##0.00'
            pw.cell(row=pnr, column=5).number_format = '"$"#,##0.00'
            pw.cell(row=pnr, column=7).number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'

            p_status = pw.cell(row=pnr, column=6)
            if bet["status"] == "Won":
                p_status.font = dfont(size=9, bold=True, color=GREEN_WIN)
            elif bet["status"] == "Lost":
                p_status.font = dfont(size=9, bold=True, color=RED_LOSS)

            parlay_legs_added += 1

    wb.save(path)
    return {"added": added, "skipped": skipped, "parlays": parlay_legs_added, "total_parsed": len(bets)}


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python parse_bovada_paste.py <paste_file.txt>")
        print("  Or pipe: pbpaste | python parse_bovada_paste.py -")
        sys.exit(1)

    source = sys.argv[1]
    if source == "-":
        raw = sys.stdin.read()
    else:
        with open(source, "r") as f:
            raw = f.read()

    bets = parse_bovada_text(raw)

    print(f"Parsed {len(bets)} bets from Bovada paste:")
    for b in bets:
        status_icon = "✅" if b["status"] == "Won" else "❌" if b["status"] == "Lost" else "➖"
        print(f"  {status_icon} {b['settled_date']} | {b['bet_type']:15s} | "
              f"${b['risk']:.2f} @ {b['odds']:>5s} | {b['status']:4s} | "
              f"${b['win_loss']:+.2f} | {b['description'][:60]}")

    if bets:
        result = write_bets_to_tracker(bets)
        print(f"\nTracker updated: {result['added']} added, "
              f"{result['skipped']} skipped (dupes), "
              f"{result['parlays']} parlays logged")
    else:
        print("No bets parsed.")


if __name__ == "__main__":
    main()
