"""
conftest.py — shared pytest fixtures.

Generates a minimal Betting_Tracker.xlsx stand-in so server/api tests
can run against a known state without touching the real file.
"""

import os
import sys
from datetime import datetime

import pytest
import openpyxl

# Make the project root importable (so tests can `import server`, etc.)
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)


# ── Fixture tracker workbook layout ──────────────────────────────────────────
# Mirrors the real Betting_Tracker.xlsx schema (first two rows reserved for
# headers; server.py expects data starting at row 3).
BET_HISTORY_HEADERS = [
    "Settled Date", "Transaction", "Sport", "Bet Type",
    "Teams/Event", "Line", "Odds", "Risk", "To Win",
    "Status", "P/L", "Notes", "Source",
]

OPEN_BETS_HEADERS = [
    "Game Start", "Source", "Sport", "Bet Type", "Teams/Event",
    "Line", "Odds", "Risk", "To Win", "Status", "Notes",
]

SAMPLE_SETTLED = [
    # A winning NBA bet at -110
    ["2026-03-24", "600000001", "NBA", "Straight",
     "Lakers vs Warriors", "-4.5", -110, 50, 45.45,
     "Won", 45.45, "", "Locks"],
    # A losing NFL bet
    ["2026-03-23", "600000002", "NFL", "Straight",
     "Chiefs vs Bills", "-3", -115, 100, 86.96,
     "Lost", -100, "", "Locks"],
    # A winning 2-leg parlay
    ["2026-03-22", "600000003", "NBA", "Parlay",
     "Lakers ML + Warriors -3", "", 260, 25, 65,
     "Won", 65, "2-leg parlay", "Bovada"],
]

SAMPLE_OPEN = [
    ["2026-04-21 00:00 UTC", "Locks25", "NBA", "Straight",
     "Celtics vs Heat", "-6.5", -110, 50, 45.45, "Pending", ""],
]


@pytest.fixture
def sample_tracker_path(tmp_path):
    """Build a self-contained fixture xlsx in a pytest tmp dir.
    Returns the absolute path to the file."""
    path = tmp_path / "Betting_Tracker.xlsx"
    wb = openpyxl.Workbook()

    # Bet History — server reads min_row=4, so we keep rows 1-3 as banner/headers/spacer
    ws = wb.active
    ws.title = "Bet History"
    ws.append([""] * len(BET_HISTORY_HEADERS))  # row 1: banner
    ws.append(BET_HISTORY_HEADERS)               # row 2: header labels
    ws.append([""] * len(BET_HISTORY_HEADERS))  # row 3: spacer (real file reserves row 3 too)
    for row in SAMPLE_SETTLED:
        ws.append(row)

    # Open Bets — same layout
    ws2 = wb.create_sheet("Open Bets")
    ws2.append([""] * len(OPEN_BETS_HEADERS))
    ws2.append(OPEN_BETS_HEADERS)
    ws2.append([""] * len(OPEN_BETS_HEADERS))
    for row in SAMPLE_OPEN:
        ws2.append(row)

    # Dashboard sheet (server.py reads it but tolerates missing keys)
    wb.create_sheet("Dashboard")

    wb.save(path)
    return str(path)


@pytest.fixture
def sample_tracker_bytes(sample_tracker_path):
    """Return the raw bytes of the fixture xlsx (for tests that need them)."""
    with open(sample_tracker_path, "rb") as f:
        return f.read()
