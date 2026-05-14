"""
server.py — Betting Tracker data bridge
─────────────────────────────────────────
Reads Betting_Tracker.xlsx and serves it as JSON to the HTML dashboard.
Also triggers the Locks25 / Bovada scrape scripts on demand.

Run once in a terminal, leave it running:

    pip install flask flask-cors openpyxl requests
    python server.py

The dashboard at betting-tracker.html auto-connects on port 5001.

Endpoints:
    GET  /api/bets              → all settled bets from "Bet History" sheet
    GET  /api/open-bets         → all open bets from "Open Bets" sheet
    GET  /api/futures-odds      → current championship odds from ESPN (no key needed)
    POST /api/refresh/locks25   → run refresh_locks25.py, return updated data
    POST /api/refresh/bovada    → run refresh_bovada.py, return updated data
    GET  /api/status            → health-check
"""

import os, sys, json, re, subprocess, time
from datetime import datetime, timedelta
import openpyxl
import requests
from flask import Flask, jsonify, request
from flask_cors import CORS

app = Flask(__name__, static_folder='.', static_url_path='/static')
CORS(app)  # allow file:// origin from the HTML dashboard


@app.route('/')
def serve_dashboard():
    """Serve the dashboard HTML — allows opening via http://localhost:5001/ instead of file://"""
    from flask import send_from_directory
    return send_from_directory(SCRIPT_DIR, 'betting-tracker.html')


@app.route('/css/<path:filename>')
def serve_css(filename):
    return send_from_directory(os.path.join(SCRIPT_DIR, 'css'), filename)


@app.route('/js/<path:filename>')
def serve_js(filename):
    return send_from_directory(os.path.join(SCRIPT_DIR, 'js'), filename)

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
LOCKS_SCRIPT = os.path.join(SCRIPT_DIR, "refresh_locks25.py")
BOVADA_SCRIPT = os.path.join(SCRIPT_DIR, "refresh_bovada.py")
ODDS_STATE_FILE = os.path.join(SCRIPT_DIR, "odds_api_state.json")

# ── multi-user support ────────────────────────────────────────────────────────
# Each user gets their own xlsx workbook. Scrapers stay pinned to Thomas.
ALLOWED_USERS = ['Thomas', 'Andrew', 'Rudger', 'Tyler', 'baby']
DEFAULT_USER  = 'Thomas'
SCRAPER_USER  = 'Thomas'   # Locks25/Bovada credentials belong to Thomas

def tracker_path_for(user):
    """Resolve the xlsx path for a given user. Falls back to default if unknown."""
    if not user or user not in ALLOWED_USERS:
        user = DEFAULT_USER
    return os.path.join(SCRIPT_DIR, f"Betting_Tracker_{user}.xlsx")

def _get_user(req):
    """Pull the user from a Flask request's ?user= param, validated against the allowlist."""
    u = (req.args.get('user') or '').strip()
    if u in ALLOWED_USERS:
        return u
    return DEFAULT_USER

# Back-compat shim so any stray reference still resolves (preflight, etc.).
TRACKER_PATH = tracker_path_for(DEFAULT_USER)

# ── scraper exit-code labels ─────────────────────────────────────────────────
# Mirrors scraper_common.py. Surfaced via /api/refresh/* so the dashboard can
# show a readable status badge instead of a raw int.
SCRAPER_EXIT_LABELS = {
    0: ("success",  "Success"),
    1: ("auth",     "Auth failed — check credentials in .env"),
    2: ("scrape",   "Scrape failed — site layout may have changed"),
    3: ("browser",  "Chrome not found — check CHROME_BINARY"),
    4: ("excel",    "Excel file is open — close Betting_Tracker.xlsx"),
    5: ("budget",   "Daily API budget exceeded — resets at midnight"),
}

def _exit_info(code):
    """Return {code, slug, label} for a subprocess exit code.
    Unknown codes get slug='unknown' so the dashboard still renders a badge."""
    slug, label = SCRAPER_EXIT_LABELS.get(code, ("unknown", f"Unknown error (exit {code})"))
    return {"code": code, "slug": slug, "label": label, "ok": code == 0}

# ── sport normalisation ───────────────────────────────────────────────────────
SPORT_MAP = {
    'cbb': 'NCAAMB', 'cbb live': 'NCAAMB',
    'nba': 'NBA',    'nba live': 'NBA',
    'nfl': 'NFL',    'nfl live': 'NFL',
    'soccer': 'Soccer',
    'ncaamb': 'NCAAMB', 'ncaawb': 'NCAAWB',
}

def normalise_sport(raw):
    if not raw:
        return 'Other'
    return SPORT_MAP.get(raw.strip().lower(), raw.strip())

# ── bet type inference ─────────────────────────────────────────────────────────
def infer_type(bet_type_col, line_col, teams_col):
    """Map Excel bet-type/line columns → dashboard type string.

    Recognised types: parlay, future, prop, total, moneyline, spread, straight.
    Teaser counts as parlay (multi-leg correlated bet). Player/event props are
    detected via 'prop' keyword in the bet_type or teams columns.
    """
    bt = (bet_type_col or '').lower()
    ln = (line_col or '').lower()
    tm = (teams_col or '').lower()

    if 'parlay' in bt or 'parlay' in ln or 'parlay' in tm or 'teaser' in bt:
        return 'parlay'
    # Futures: detect via Line column ("Future"), Teams keywords, or known league/cup names
    if ('future' in ln or 'future' in tm or 'odds to win' in tm or 'championship' in tm
            or 'premier league' in tm or 'champions league' in tm
            or 'la liga' in tm or 'serie a' in tm or 'bundesliga' in tm
            or 'world cup' in tm or 'euros' in tm or 'copa america' in tm):
        return 'future'
    if 'prop' in bt or 'prop' in tm:
        return 'prop'
    if 'over' in ln or 'under' in ln or 'total' in ln or '1h' in ln.replace('-',''):
        return 'total'
    if 'ml' in ln or 'moneyline' in ln:
        return 'moneyline'
    if 'ats' in ln or re.search(r'[+-]\d+\.?\d*', ln):
        return 'spread'
    return 'straight'

# ── odds parsing ──────────────────────────────────────────────────────────────
def parse_odds(raw):
    if raw is None:
        return 0
    s = str(raw).strip().replace(' ', '')
    # Handle implied odds like "-112 implied" or "+103 implied"
    m = re.search(r'([+-]\d+)', s)
    if m:
        try:
            return int(m.group(1))
        except:
            pass
    return 0

# ── date helpers ──────────────────────────────────────────────────────────────
MONTH_ABBR = {
    'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
    'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12
}

def parse_date_str(s):
    """Parse 'Mar-22-2026' or 'Mar-22-2026 07:00 PM' → ISO 8601 string.
    Returns None on unparseable input instead of silently substituting utcnow()."""
    if not s:
        return None
    s = str(s).strip()
    # Full datetime: "Mar-26-2026 07:00 PM"
    m = re.match(r'(\w+)-(\d+)-(\d+)\s+(\d+):(\d+)\s*(AM|PM)', s, re.IGNORECASE)
    if m:
        mon, day, yr, hr, mn, ampm = m.groups()
        hr = int(hr)
        if ampm.upper() == 'PM' and hr != 12:
            hr += 12
        if ampm.upper() == 'AM' and hr == 12:
            hr = 0
        try:
            dt = datetime(int(yr), MONTH_ABBR.get(mon[:3], 1), int(day), hr, int(mn))
            return dt.isoformat() + '.000Z'
        except Exception:
            pass
    # Date only: "Mar-22-2026"
    m = re.match(r'(\w+)-(\d+)-(\d+)', s)
    if m:
        mon, day, yr = m.groups()
        try:
            dt = datetime(int(yr), MONTH_ABBR.get(mon[:3], 1), int(day), 20, 0)
            return dt.isoformat() + '.000Z'
        except Exception:
            pass
    print(f"[WARN] Could not parse date: '{s}' — returning None")
    return None

def format_game_time(s):
    """'Mar-26-2026 07:00 PM' → '3/26/26 7:00 PM'
       'Mar-22-2026'          → '3/22/26 8:00 PM'  (default 8 PM for date-only)"""
    if not s:
        return ''
    s = str(s).strip()
    # Full datetime: "Mar-26-2026 07:00 PM"
    m = re.match(r'(\w+)-(\d+)-(\d+)\s+(.*)', s)
    if m:
        mon, day, yr, time_part = m.groups()
        mon_num = MONTH_ABBR.get(mon[:3], 1)
        yr_short = str(yr)[-2:]
        # Clean up time (remove leading zero from hour)
        tp = re.sub(r'^0(\d)', r'\1', time_part)
        return f"{mon_num}/{int(day)}/{yr_short} {tp}"
    # Date only: "Mar-22-2026" — add default 8:00 PM (typical game evening)
    m = re.match(r'(\w+)-(\d+)-(\d+)$', s)
    if m:
        mon, day, yr = m.groups()
        mon_num = MONTH_ABBR.get(mon[:3], 1)
        yr_short = str(yr)[-2:]
        return f"{mon_num}/{int(day)}/{yr_short} 8:00 PM"
    return s

# ── Excel caching (per-user, dirty flag with file modification time) ─────────

# Keyed by user name. Each entry: {'settled': [...]|None, 'open': [...]|None, 'mtime': float}
_xlsx_caches = {}

def _get_cache(user):
    """Return (and lazily create) the cache slot for a user."""
    if user not in _xlsx_caches:
        _xlsx_caches[user] = {'settled': None, 'open': None, 'mtime': 0}
    return _xlsx_caches[user]

def _xlsx_is_stale(user):
    """True when the user's xlsx file has changed since the last read."""
    cache = _get_cache(user)
    try:
        mtime = os.path.getmtime(tracker_path_for(user))
        return mtime > cache['mtime']
    except OSError:
        return True

def _invalidate_xlsx_cache(user=None):
    """Drop the cache for a user (or all users if user is None)."""
    if user is None:
        for c in _xlsx_caches.values():
            c['mtime'] = 0
    else:
        _get_cache(user)['mtime'] = 0

# ── Excel readers ─────────────────────────────────────────────────────────────

def read_settled_bets(user=DEFAULT_USER):
    """Read all rows from 'Bet History' sheet for a user, return dashboard-format dicts.
    Uses file mtime-based caching to avoid re-parsing on every request."""
    cache = _get_cache(user)
    path  = tracker_path_for(user)
    if not _xlsx_is_stale(user) and cache['settled'] is not None:
        return cache['settled']

    bets = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb['Bet History']
        # enumerate so we can include the 1-based Excel row index in each bet —
        # used by the dashboard's Edit feature when txId lookup is ambiguous.
        for excel_row, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            settled_date = row[0]
            tx_id        = row[1]
            sport        = row[2]
            bet_type     = row[3]
            teams        = row[4]
            line         = row[5]
            odds_raw     = row[6]
            risk         = row[7]
            to_win       = row[8]
            status       = row[9]
            win_loss     = row[10]
            notes        = row[11]
            source_col   = row[12] if len(row) > 12 else None

            if not tx_id:   # skip empty / filler rows
                continue

            result = ''
            if str(status or '').lower() == 'won':
                result = 'W'
            elif str(status or '').lower() == 'lost':
                result = 'L'
            elif str(status or '').lower() in ('push', 'tie'):
                result = 'P'

            # Build a readable pick string: "Teams / Event  Line (+Odds)"
            pick_parts = []
            if teams:
                pick_parts.append(str(teams).strip())
            if line:
                pick_parts.append(str(line).strip())
            odds_num = parse_odds(odds_raw)
            if odds_num:
                sign = '+' if odds_num > 0 else ''
                pick_parts.append(f'({sign}{odds_num})')

            bets.append({
                'id':          str(tx_id),
                'txId':        str(tx_id),
                'sport':       normalise_sport(sport),
                'type':        infer_type(bet_type, line, teams),
                'matchup':     str(teams or ''),
                'pick':        '  '.join(pick_parts),
                'odds':        odds_num,
                'stake':       float(risk or 0),
                'toWin':       float(to_win or 0),
                'settled':     result in ('W', 'L', 'P'),
                'result':      result,
                'winLoss':     float(win_loss or 0),
                'notes':       str(notes or ''),
                'settledDate': parse_date_str(settled_date),
                'addedDate':   parse_date_str(settled_date),
                'gameTime':    format_game_time(settled_date),
                'source':      str(source_col or 'Locks'),
                'excelRow':    excel_row,
                'excelSheet':  'history',
            })
        wb.close()

        # Commit the mtime so a standalone /api/bets call doesn't re-read on every
        # request. read_open_bets() snapshots current_mtime before calling us and
        # overwrites cache['mtime'] at its end, so there is no double-update hazard.
        cache['mtime']   = os.path.getmtime(path)
        cache['settled'] = bets
    except (FileNotFoundError, PermissionError):
        # Let the /api endpoint map these to XLSX_MISSING / XLSX_LOCKED.
        cache['settled'] = None
        cache['mtime'] = 0
        raise
    except Exception as e:
        print(f"Error reading settled bets from Excel ({user}): {e}")
        cache['settled'] = []
        cache['mtime'] = 0
        return []

    return bets


def read_open_bets(user=DEFAULT_USER):
    """Read all rows from 'Open Bets' sheet for a user, return dashboard-format dicts.
    Uses file mtime-based caching to avoid re-parsing on every request."""
    cache = _get_cache(user)
    path  = tracker_path_for(user)
    if not _xlsx_is_stale(user) and cache['open'] is not None:
        return cache['open']

    # If settled cache is also stale, read it first so both caches refresh together
    if cache['settled'] is None or _xlsx_is_stale(user):
        read_settled_bets(user)

    bets = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb['Open Bets']
        for excel_row, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            game_time = row[0]
            source    = row[1]
            sport     = row[2]
            bet_type  = row[3]
            teams     = row[4]
            line      = row[5]
            odds_raw  = row[6]
            risk      = row[7]
            to_win    = row[8]
            status    = row[9]
            notes     = row[10]

            if not teams:
                continue

            # Extract ticket ID from notes like "Placed: Feb-05-2026  |  Ticket: 608309226"
            tx_id = ''
            if notes:
                m = re.search(r'Ticket:\s*(\d+)', str(notes))
                if m:
                    tx_id = m.group(1)

            pick_parts = []
            if teams:
                pick_parts.append(str(teams).strip())
            if line:
                pick_parts.append(str(line).strip())
            odds_num = parse_odds(odds_raw)
            if odds_num:
                sign = '+' if odds_num > 0 else ''
                pick_parts.append(f'({sign}{odds_num})')

            bets.append({
                'id':       tx_id or f'open_{len(bets)}',
                'txId':     tx_id,
                'sport':    normalise_sport(sport),
                'type':     infer_type(bet_type, line, teams),
                'matchup':  str(teams or ''),
                'pick':     '  '.join(pick_parts),
                'odds':     odds_num,
                'stake':    float(risk or 0),
                'toWin':    float(to_win or 0),
                'settled':  False,
                'result':   '',
                'notes':    str(notes or ''),
                'gameTime': format_game_time(game_time),
                'addedDate': parse_date_str(game_time),
                'source':   str(source or 'Locks25'),
                'excelRow': excel_row,
                'excelSheet': 'open',
            })
        wb.close()

        # Both settled and open are now fresh — update mtime so cache is valid
        cache['open'] = bets
        cache['mtime'] = os.path.getmtime(path)
    except (FileNotFoundError, PermissionError):
        cache['open'] = None
        cache['mtime'] = 0
        raise
    except Exception as e:
        print(f"Error reading open bets from Excel ({user}): {e}")
        cache['open'] = []
        cache['mtime'] = 0
        return []

    return bets

# ── routes ────────────────────────────────────────────────────────────────────

@app.route('/api/status', methods=['GET'])
def status():
    """Health check with cache status information."""
    user = _get_user(request)
    cache = _get_cache(user)
    xlsx_cache_age = None
    try:
        if cache['mtime'] > 0:
            xlsx_cache_age = round(time.time() - cache['mtime'], 1)
    except Exception:
        pass

    # _futures_cache is a per-sports-key dict: {key: {'data': ..., 'timestamp': float}}
    futures_entries = list(_futures_cache.values())
    futures_cached = len(futures_entries) > 0
    futures_cache_age = None
    if futures_cached:
        # Report the most recent cache entry's age
        newest = max(futures_entries, key=lambda e: e['timestamp'])
        futures_cache_age = round(time.time() - newest['timestamp'], 1)

    return jsonify({
        'ok': True,
        'time': datetime.utcnow().isoformat(),
        'user': user,
        'users': ALLOWED_USERS,
        'cache': {
            'xlsx_cached': cache['settled'] is not None,
            'xlsx_cache_age_seconds': xlsx_cache_age,
            'futures_cached': futures_cached,
            'futures_cache_age_seconds': futures_cache_age,
            'futures_cache_ttl_seconds': _FUTURES_CACHE_TTL,
            'futures_cache_keys': list(_futures_cache.keys()),
        }
    })


def _excel_error_payload(e, path=None):
    """Map openpyxl / OS errors to stable error codes the dashboard can react to."""
    msg = str(e)
    if isinstance(e, PermissionError) or 'Permission denied' in msg or 'being used by another process' in msg:
        return {'ok': False, 'error': 'Excel workbook is locked (close it in Excel and retry).', 'code': 'XLSX_LOCKED'}, 503
    if isinstance(e, FileNotFoundError):
        return {'ok': False, 'error': f'Workbook not found at {path or "?"}.', 'code': 'XLSX_MISSING'}, 503
    return {'ok': False, 'error': msg, 'code': 'XLSX_ERROR'}, 500


@app.route('/api/bets', methods=['GET'])
def get_bets():
    """Return all settled bets from the Excel for ?user= (default Thomas)."""
    user = _get_user(request)
    try:
        bets = read_settled_bets(user)
        return jsonify({'ok': True, 'user': user, 'count': len(bets), 'bets': bets})
    except Exception as e:
        payload, status_code = _excel_error_payload(e, tracker_path_for(user))
        return jsonify(payload), status_code


@app.route('/api/open-bets', methods=['GET'])
def get_open_bets():
    """Return all open bets from the Excel for ?user= (default Thomas)."""
    user = _get_user(request)
    try:
        bets = read_open_bets(user)
        return jsonify({'ok': True, 'user': user, 'count': len(bets), 'bets': bets})
    except Exception as e:
        payload, status_code = _excel_error_payload(e, tracker_path_for(user))
        return jsonify(payload), status_code


# ── manual edit (single-row update) ──────────────────────────────────────────
# Field → (sheet column index) for each sheet. Only fields listed here can be edited.
# Bet History: A=settledDate B=txId C=sport D=betType E=teams F=line G=odds
#              H=risk I=toWin J=status K=winLoss L=notes M=source
# Open Bets:   A=gameTime B=source C=sport D=betType E=teams F=line G=odds
#              H=risk I=toWin J=status K=notes
HISTORY_FIELD_COLS = {
    'settledDate': 1, 'txId': 2, 'sport': 3, 'betType': 4, 'teams': 5,
    'line': 6, 'odds': 7, 'stake': 8, 'toWin': 9, 'status': 10,
    'winLoss': 11, 'notes': 12, 'source': 13,
}
OPEN_FIELD_COLS = {
    'gameTime': 1, 'source': 2, 'sport': 3, 'betType': 4, 'teams': 5,
    'line': 6, 'odds': 7, 'stake': 8, 'toWin': 9, 'status': 10, 'notes': 11,
}

def _coerce_cell_value(field, raw):
    """Normalise incoming JSON values to the right Excel cell type."""
    if raw is None or raw == '':
        return None
    if field in ('stake', 'toWin', 'winLoss'):
        try:
            return float(raw)
        except (TypeError, ValueError):
            return None
    if field == 'odds':
        # Accept '+150', '-110', 150, '150' — store as the original string format
        # the readers expect ("-110" or "+150"). Also accept ints.
        s = str(raw).strip()
        if s and s[0] not in '+-' and s.lstrip('-').isdigit():
            n = int(s)
            return f'+{n}' if n > 0 else str(n)
        return s
    return str(raw)

def _find_row_index(ws, sheet_kind, tx_id, row_key):
    """Locate the 1-based Excel row for a bet on a given sheet.
    Priority: txId match (Bet History col B, Open Bets ticket-in-notes regex),
    then explicit row_key (1-based)."""
    if sheet_kind == 'history':
        if tx_id:
            for r in range(4, ws.max_row + 1):
                v = ws.cell(row=r, column=2).value
                if v is not None and str(v).strip() == str(tx_id).strip():
                    return r
        if row_key:
            try:
                rk = int(row_key)
                if rk >= 4:
                    return rk
            except (TypeError, ValueError):
                pass
        return None
    # open bets — match Ticket: <txId> in notes (col K = index 11)
    if tx_id:
        pat = re.compile(r'Ticket:\s*' + re.escape(str(tx_id).strip()))
        for r in range(4, ws.max_row + 1):
            notes = ws.cell(row=r, column=11).value
            if notes and pat.search(str(notes)):
                return r
    if row_key:
        try:
            rk = int(row_key)
            if rk >= 4:
                return rk
        except (TypeError, ValueError):
            pass
    return None


@app.route('/api/bets/update', methods=['POST'])
def update_bet():
    """Manually edit a single row in Bet History or Open Bets. Used by the
    dashboard's Edit button when the parser captured something wrong.

    Body:
        {
            "user":   "Thomas",                # optional, defaults to ?user= or DEFAULT_USER
            "sheet":  "history" | "open",
            "txId":   "608309226",             # preferred lookup
            "rowKey": 12,                      # fallback: 1-based Excel row index
            "fields": { "sport": "NBA", "odds": "-110", ... }
        }

    Returns the refreshed bet object plus updated counts so the dashboard can
    update localStorage in place without a full sync."""
    payload = request.get_json(silent=True) or {}
    user    = payload.get('user') or _get_user(request)
    if user not in ALLOWED_USERS:
        user = DEFAULT_USER
    sheet_kind = (payload.get('sheet') or '').lower()
    if sheet_kind not in ('history', 'open'):
        return jsonify({'ok': False, 'error': "sheet must be 'history' or 'open'", 'code': 'BAD_SHEET'}), 400
    fields = payload.get('fields') or {}
    if not isinstance(fields, dict) or not fields:
        return jsonify({'ok': False, 'error': 'fields object is required', 'code': 'NO_FIELDS'}), 400

    tx_id   = payload.get('txId')
    row_key = payload.get('rowKey')

    col_map = HISTORY_FIELD_COLS if sheet_kind == 'history' else OPEN_FIELD_COLS
    sheet_name = 'Bet History' if sheet_kind == 'history' else 'Open Bets'

    # Reject unknown fields up front so a typo doesn't silently no-op
    unknown = [k for k in fields.keys() if k not in col_map]
    if unknown:
        return jsonify({'ok': False, 'error': f'unknown fields for {sheet_kind}: {unknown}',
                        'code': 'UNKNOWN_FIELDS', 'allowed': list(col_map.keys())}), 400

    path = tracker_path_for(user)
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return jsonify({'ok': False, 'error': f'sheet "{sheet_name}" not found',
                            'code': 'SHEET_MISSING'}), 404
        ws = wb[sheet_name]
        row = _find_row_index(ws, sheet_kind, tx_id, row_key)
        if not row:
            wb.close()
            return jsonify({'ok': False, 'error': 'bet not found by txId or rowKey',
                            'code': 'ROW_NOT_FOUND'}), 404

        # Snapshot before/after for the audit trail in the response
        before = {}
        after  = {}
        for field, value in fields.items():
            col = col_map[field]
            before[field] = ws.cell(row=row, column=col).value
            new_val = _coerce_cell_value(field, value)
            ws.cell(row=row, column=col).value = new_val
            after[field] = new_val

        wb.save(path)
        wb.close()
    except PermissionError as e:
        payload, status_code = _excel_error_payload(e, path)
        return jsonify(payload), status_code
    except FileNotFoundError as e:
        payload, status_code = _excel_error_payload(e, path)
        return jsonify(payload), status_code
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'code': 'XLSX_ERROR'}), 500

    # Force a re-read so the next GET reflects the edit
    _invalidate_xlsx_cache(user)
    settled = read_settled_bets(user)
    open_b  = read_open_bets(user)

    # Find and return the just-edited bet so the client can patch it in-place
    updated = None
    pool = settled if sheet_kind == 'history' else open_b
    if tx_id:
        for b in pool:
            if str(b.get('txId') or '') == str(tx_id):
                updated = b
                break
    return jsonify({
        'ok': True,
        'user': user,
        'sheet': sheet_kind,
        'row': row,
        'before': before,
        'after': after,
        'bet': updated,
        'settled_count': len(settled),
        'open_count': len(open_b),
    })


@app.route('/api/refresh/locks25', methods=['POST'])
def refresh_locks25():
    """Run refresh_locks25.py, then return full updated dataset.
    Always operates on Thomas's xlsx — scrapers are pinned to Thomas's accounts.
    Surfaces the scraper exit code via the 'status' field so the dashboard
    can show a readable badge (auth / scrape / browser / excel / budget)."""
    status = None
    try:
        if os.path.exists(LOCKS_SCRIPT):
            result = subprocess.run(
                [sys.executable, LOCKS_SCRIPT],
                capture_output=True, text=True, timeout=120,
                cwd=SCRIPT_DIR
            )
            status = _exit_info(result.returncode)
            if result.returncode != 0:
                print(f"refresh_locks25.py exited {result.returncode} ({status['slug']}):", result.stderr[-500:])
        else:
            status = {"code": -1, "slug": "missing", "label": "Scraper script not found", "ok": False}

        # Scrapers write to Thomas's xlsx — invalidate his cache only
        _invalidate_xlsx_cache(SCRAPER_USER)

        settled  = read_settled_bets(SCRAPER_USER)
        open_b   = read_open_bets(SCRAPER_USER)
        return jsonify({
            'ok': True,
            'user': SCRAPER_USER,
            'status': status,
            'settled_count': len(settled),
            'open_count':    len(open_b),
            'bets':      settled,
            'open_bets': open_b,
        })
    except Exception as e:
        return jsonify({'ok': False, 'status': status, 'error': str(e)}), 500


@app.route('/api/refresh/bovada', methods=['POST'])
def refresh_bovada():
    """Run refresh_bovada.py, then return full updated dataset.
    Always operates on Thomas's xlsx — scrapers are pinned to Thomas's accounts.
    See refresh_locks25 for the status-badge contract."""
    status = None
    try:
        if os.path.exists(BOVADA_SCRIPT):
            result = subprocess.run(
                [sys.executable, BOVADA_SCRIPT],
                capture_output=True, text=True, timeout=120,
                cwd=SCRIPT_DIR
            )
            status = _exit_info(result.returncode)
            if result.returncode != 0:
                print(f"refresh_bovada.py exited {result.returncode} ({status['slug']}):", result.stderr[-500:])
        else:
            status = {"code": -1, "slug": "missing", "label": "Scraper script not found", "ok": False}

        # Scrapers write to Thomas's xlsx — invalidate his cache only
        _invalidate_xlsx_cache(SCRAPER_USER)

        settled  = read_settled_bets(SCRAPER_USER)
        open_b   = read_open_bets(SCRAPER_USER)
        return jsonify({
            'ok': True,
            'user': SCRAPER_USER,
            'status': status,
            'settled_count': len(settled),
            'open_count':    len(open_b),
            'bets':      settled,
            'open_bets': open_b,
        })
    except Exception as e:
        return jsonify({'ok': False, 'status': status, 'error': str(e)}), 500


# ── futures odds ──────────────────────────────────────────────────────────────
# Primary: Bovada public API (free, no auth, returns JSON with American odds)
# Fallback: ESPN (free, no auth, less reliable)
# Odds history persisted to odds_history.json for line movement tracking.

ODDS_HISTORY_FILE   = os.path.join(SCRIPT_DIR, 'odds_history.json')
FUTURES_CONFIG_FILE = os.path.join(SCRIPT_DIR, 'futures_config.json')


def _load_futures_config():
    """Load persisted futures config (API key, sports) from disk."""
    if os.path.exists(FUTURES_CONFIG_FILE):
        try:
            with open(FUTURES_CONFIG_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_futures_config(config):
    """Persist futures config to disk so the scheduled refresh can use the key."""
    try:
        with open(FUTURES_CONFIG_FILE, 'w') as f:
            json.dump(config, f, indent=2)
    except Exception as e:
        print(f"Warning: could not save futures_config.json: {e}")


# Futures odds cache — keyed by sorted sports string so nba vs nba+ncaamb are cached separately.
# TTL is 6 hours to stay comfortably within The Odds API free tier (500 req/month).
_futures_cache = {}   # { cache_key: {'data': {...}, 'timestamp': float} }
_FUTURES_CACHE_TTL = 21600  # 6 hours

def _futures_cache_key(sports):
    """Stable key from any sports list — order-independent."""
    return ','.join(sorted(set(sports)))

def _futures_cache_valid(cache_key):
    """True when the sports-specific cache entry is still fresh."""
    entry = _futures_cache.get(cache_key)
    return entry is not None and (time.time() - entry['timestamp'] < _FUTURES_CACHE_TTL)

def _futures_cache_get(cache_key):
    return _futures_cache.get(cache_key, {}).get('data')

def _futures_cache_set(cache_key, data):
    _futures_cache[cache_key] = {'data': data, 'timestamp': time.time()}

# Bovada public API paths for championship/futures markets
BOVADA_FUTURES = {
    'nba':    '/basketball/nba-championship',
    'ncaamb': '/basketball/college-basketball/college-basketball-futures',
    'cbb':    '/basketball/college-basketball/college-basketball-futures',
    'nfl':    '/football/nfl-specials',
    'mlb':    '/baseball/mlb-season-specials',
    'nhl':    '/hockey/nhl-specials',
}
BOVADA_BASE = 'https://www.bovada.lv/services/sports/event/v2/events/A/description'

# ESPN fallback endpoints
SPORT_ENDPOINTS = {
    'nba':        ('basketball', 'nba'),
    'ncaamb':     ('basketball', 'mens-college-basketball'),
    'cbb':        ('basketball', 'mens-college-basketball'),
    'ncaawb':     ('basketball', 'womens-college-basketball'),
    'nfl':        ('football',   'nfl'),
    'mlb':        ('baseball',   'mlb'),
    'nhl':        ('hockey',     'nhl'),
    'soccer_ucl': ('soccer',     'uefa.champions_league'),
    'soccer_epl': ('soccer',     'eng.1'),
    'soccer':     ('soccer',     'uefa.champions_league'),
}

# The Odds API — free tier (500 req/month), covers NBA/NCAAM/Soccer/NFL/MLB/NHL
# Sign up at https://the-odds-api.com to get a free key, then enter it in Settings.
THE_ODDS_API_BASE = 'https://api.the-odds-api.com/v4/sports'
THE_ODDS_API_SPORTS = {
    'nba':        'basketball_nba_championship_winner',
    'ncaamb':     'basketball_ncaab_championship_winner',
    'cbb':        'basketball_ncaab_championship_winner',
    'nfl':        'americanfootball_nfl_super_bowl_winner',
    'mlb':        'baseball_mlb_world_series_winner',
    'nhl':        'icehockey_nhl_championship_winner',
    'soccer_ucl': 'soccer_uefa_champs_league_winner',
    'soccer_epl': 'soccer_epl_winner',
    'soccer':     'soccer_uefa_champs_league_winner',
}

_FETCH_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                  'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Accept': 'application/json',
}


def _load_odds_history():
    """Load odds history from JSON file. Structure:
    { "team_name_lower": [ { "odds": -150, "bookmaker": "Bovada", "ts": "2026-03-24T12:00:00Z" }, ... ] }
    """
    if os.path.exists(ODDS_HISTORY_FILE):
        try:
            with open(ODDS_HISTORY_FILE) as f:
                return json.load(f)
        except:
            pass
    return {}


def _save_odds_history(history):
    """Persist odds history to disk."""
    try:
        with open(ODDS_HISTORY_FILE, 'w') as f:
            json.dump(history, f, indent=1)
    except Exception as e:
        print(f"Failed to save odds history: {e}")


def _append_odds_to_history(current_odds):
    """Append today's odds snapshot to the history file.
    Only stores one entry per team per calendar day to keep the file lean."""
    if not current_odds:
        return
    history = _load_odds_history()
    today = datetime.utcnow().strftime('%Y-%m-%d')
    ts = datetime.utcnow().isoformat() + 'Z'

    for team, info in current_odds.items():
        if team not in history:
            history[team] = []
        entries = history[team]
        already_today = False
        for e in entries:
            if e.get('ts', '').startswith(today):
                e['odds'] = info['odds']
                e['bookmaker'] = info.get('bookmaker', '')
                e['ts'] = ts
                already_today = True
                break
        if not already_today:
            entries.append({
                'odds': info['odds'],
                'bookmaker': info.get('bookmaker', ''),
                'ts': ts,
            })
        if len(entries) > 90:
            history[team] = entries[-90:]

    _save_odds_history(history)


def _parse_bovada_american_odds(price_obj):
    """Extract American odds integer from Bovada price object."""
    if not price_obj:
        return None
    # Bovada returns {'american': '+350', 'decimal': '4.50', ...}
    am = str(price_obj.get('american', '')).strip()
    if am:
        try:
            # Python's int() handles both "+350" and "-110", so just parse directly
            return int(am)
        except (ValueError, TypeError):
            pass
    # Fallback: convert decimal to American
    dec = price_obj.get('decimal')
    if dec:
        try:
            d = float(dec)
            if d >= 2.0:
                return int(round((d - 1) * 100))
            elif d > 1.0:
                return int(round(-100 / (d - 1)))
        except:
            pass
    return None


def _fetch_the_odds_api(sports, api_key):
    """Fetch futures odds from The Odds API (free tier: 500 req/month).
    Returns {team_lower: {odds, bookmaker}}, [errors].
    Preferred bookmaker order: FanDuel → DraftKings → BetMGM → Bovada → first available."""
    if not api_key:
        return {}, []

    all_odds = {}
    errors = []
    PREFERRED_BOOKS = ['fanduel', 'draftkings', 'betmgm', 'bovada', 'williamhill_us', 'betonlineag']

    for sport in sports:
        sport_key = THE_ODDS_API_SPORTS.get(sport)
        if not sport_key:
            continue
        url = f'{THE_ODDS_API_BASE}/{sport_key}/odds/'
        params = {
            'apiKey': api_key,
            'regions': 'us',
            'markets': 'outrights',
            'oddsFormat': 'american',
        }
        try:
            r = requests.get(url, params=params, headers=_FETCH_HEADERS, timeout=12)
            remaining = r.headers.get('x-requests-remaining', '?')
            if r.status_code == 200:
                events = r.json()
                count = 0
                for event in events:
                    bookmakers = event.get('bookmakers', [])
                    # Pick the highest-priority available bookmaker
                    chosen_bm = None
                    for pref in PREFERRED_BOOKS:
                        for bm in bookmakers:
                            if bm.get('key') == pref:
                                chosen_bm = bm
                                break
                        if chosen_bm:
                            break
                    if not chosen_bm and bookmakers:
                        chosen_bm = bookmakers[0]
                    if not chosen_bm:
                        continue
                    for market in chosen_bm.get('markets', []):
                        if market.get('key') == 'outrights':
                            for outcome in market.get('outcomes', []):
                                name = outcome.get('name', '')
                                price = outcome.get('price')
                                if name and price is not None:
                                    try:
                                        all_odds[name.lower()] = {
                                            'odds': round(float(price)),
                                            'bookmaker': chosen_bm.get('title', 'The Odds API'),
                                        }
                                        count += 1
                                    except (ValueError, TypeError):
                                        pass
                print(f"  TheOddsAPI [{sport}]: {count} outcomes fetched (remaining: {remaining})")
            elif r.status_code == 401:
                errors.append(f'{sport}: The Odds API — invalid key')
            elif r.status_code == 422:
                errors.append(f'{sport}: The Odds API — sport not currently available')
            elif r.status_code == 429:
                errors.append(f'{sport}: The Odds API — monthly quota exceeded')
            else:
                errors.append(f'{sport}: The Odds API HTTP {r.status_code}')
        except Exception as exc:
            errors.append(f'{sport}: The Odds API error: {str(exc)[:100]}')

    return all_odds, errors


def _fetch_bovada(sports):
    """Fetch futures odds from Bovada's public API. No auth needed.
    Returns {team_lower: {odds, bookmaker}}."""
    all_odds = {}
    errors = []

    for sport in sports:
        path = BOVADA_FUTURES.get(sport)
        if not path:
            continue
        url = BOVADA_BASE + path
        try:
            r = requests.get(url, headers=_FETCH_HEADERS, timeout=15)
            if r.status_code == 200:
                data = r.json()
                events = data if isinstance(data, list) else [data]
                count = 0
                for ev in events:
                    for dg in ev.get('displayGroups', []):
                        desc = (dg.get('description') or '').lower()
                        # Bovada futures pages are pre-filtered, so accept all display groups
                        for mkt in dg.get('markets', []):
                            for outcome in mkt.get('outcomes', []):
                                name = outcome.get('description', '')
                                price = outcome.get('price', {})
                                odds = _parse_bovada_american_odds(price)
                                if name and odds is not None:
                                    all_odds[name.lower()] = {
                                        'odds': odds,
                                        'bookmaker': 'Bovada',
                                    }
                                    count += 1
                print(f"  Bovada [{sport}]: {count} outcomes fetched")
            elif r.status_code == 404:
                errors.append(f'{sport}: Bovada futures market not found (may be off-season)')
            else:
                errors.append(f'{sport}: Bovada HTTP {r.status_code}')
        except Exception as exc:
            errors.append(f'{sport}: Bovada error: {str(exc)[:100]}')

    return all_odds, errors


def _parse_espn_futures(data):
    """Parse ESPN futures JSON into {team_name_lower: {odds, bookmaker}}."""
    out = {}
    if not isinstance(data, dict):
        return out

    for event in data.get('events', []):
        for comp in event.get('competitions', []):
            for ctr in comp.get('competitors', []):
                name = (ctr.get('team') or {}).get('displayName') or ctr.get('displayName') or ctr.get('name', '')
                if not name:
                    continue
                odds_obj = ctr.get('odds') or {}
                ml = None
                for field in ('moneyLine', 'moneyline', 'value', 'american', 'americanOdds'):
                    if field in odds_obj:
                        try:
                            ml = int(float(odds_obj[field]))
                            break
                        except (ValueError, TypeError):
                            pass
                book = (odds_obj.get('provider') or {}).get('name') or odds_obj.get('providerName') or 'ESPN BET'
                if ml is not None:
                    out[name.lower()] = {'odds': ml, 'bookmaker': book}

    for sport_item in data.get('sports', []):
        for league in sport_item.get('leagues', []):
            for fut in league.get('futures', []):
                book = (fut.get('bookmaker') or {}).get('name', 'ESPN BET')
                for entry in fut.get('entries', []):
                    name = (
                        entry.get('competitor')
                        or (entry.get('team') or {}).get('displayName')
                        or entry.get('name', '')
                    )
                    ml = entry.get('moneyLine') or entry.get('odds') or entry.get('value')
                    if name and ml is not None:
                        try:
                            out[name.lower()] = {'odds': int(float(ml)), 'bookmaker': book}
                        except (ValueError, TypeError):
                            pass

    for item in data.get('items', data.get('futures', [])):
        name = (
            (item.get('team') or {}).get('displayName')
            or item.get('name')
            or item.get('competitor', '')
        )
        ml = item.get('moneyLine') or item.get('odds') or item.get('value')
        if name and ml is not None:
            try:
                out[name.lower()] = {'odds': int(float(ml)), 'bookmaker': item.get('bookmaker', 'ESPN BET')}
            except (ValueError, TypeError):
                pass
    return out


def _fetch_espn(sports):
    """Fetch futures odds from ESPN. Returns {team_lower: {odds, bookmaker}}."""
    all_odds = {}
    errors = []

    for sport in sports:
        pair = SPORT_ENDPOINTS.get(sport)
        if not pair:
            continue
        sport_key, league_key = pair
        endpoints = [
            f'https://site.api.espn.com/apis/site/v2/sports/{sport_key}/{league_key}/futures',
            f'https://sports.core.api.espn.com/v2/sports/{sport_key}/leagues/{league_key}/futures?limit=100',
        ]
        fetched = False
        for url in endpoints:
            try:
                r = requests.get(url, headers=_FETCH_HEADERS, timeout=12)
                if r.status_code == 200:
                    parsed = _parse_espn_futures(r.json())
                    if parsed:
                        all_odds.update(parsed)
                        fetched = True
                        print(f"  ESPN [{sport}]: {len(parsed)} teams fetched")
                        break
                    else:
                        errors.append(f'{sport}: ESPN 200 but no odds parsed')
                else:
                    errors.append(f'{sport}: ESPN HTTP {r.status_code}')
            except Exception as exc:
                errors.append(f'{sport}: ESPN error: {str(exc)[:100]}')
    return all_odds, errors


@app.route('/api/futures-odds', methods=['GET'])
def get_futures_odds():
    """
    Return current championship odds for one or more sports.
    Waterfall: The Odds API (if key provided) → Bovada → ESPN.
    Each source fills in gaps left by the prior one (per-team, not per-sport).
    Persists a daily snapshot to odds_history.json for line movement.
    Cache TTL is 6 hours per unique sports combination to stay within free API tiers.

    Query params:
        sports        comma-separated list, e.g. nba,ncaamb,soccer_ucl (default: nba)
        odds_api_key  optional The Odds API key for higher reliability + soccer coverage
    """
    sports_param = request.args.get('sports', 'nba')
    sports = [s.strip().lower() for s in sports_param.split(',') if s.strip()]
    api_key = request.args.get('odds_api_key', '').strip()

    # Persist the key to disk when the browser sends it (so the scheduler can use it too)
    cfg = _load_futures_config()
    if api_key:
        if cfg.get('odds_api_key') != api_key:
            cfg['odds_api_key'] = api_key
            cfg['sports'] = sports  # remember which sports to fetch
            _save_futures_config(cfg)
    elif not api_key:
        # Fall back to the persisted key so the endpoint still works when called without one
        api_key = cfg.get('odds_api_key', '')

    # Also update the remembered sports list whenever new sports appear
    if api_key:
        known = set(cfg.get('sports', []))
        if set(sports) - known:
            cfg['sports'] = list(known | set(sports))
            _save_futures_config(cfg)

    force_refresh = request.args.get('force_refresh', '0') in ('1', 'true', 'yes')

    # Sport-aware cache check (different sport combos cached separately)
    cache_key = _futures_cache_key(sports)
    if not force_refresh and _futures_cache_valid(cache_key):
        cached = _futures_cache_get(cache_key)
        return jsonify({
            'ok':     True,
            'odds':   cached.get('odds', {}),
            'count':  cached.get('count', 0),
            'source': cached.get('source', 'cache'),
            'errors': cached.get('errors'),
            'cached': True,
        })

    # Also check the file-based cache written by refresh_futures.py (standalone mode)
    futures_cache_file = os.path.join(SCRIPT_DIR, 'futures_cache.json')
    if not force_refresh and os.path.exists(futures_cache_file):
        try:
            with open(futures_cache_file) as f:
                file_cache = json.load(f)
            age = time.time() - file_cache.get('timestamp', 0)
            if age < _FUTURES_CACHE_TTL and file_cache.get('odds'):
                print(f'  Serving from futures_cache.json (age {age/3600:.1f}h)')
                return jsonify({
                    'ok':     True,
                    'odds':   file_cache['odds'],
                    'count':  file_cache.get('count', len(file_cache['odds'])),
                    'source': file_cache.get('source', 'scheduled_refresh'),
                    'errors': None,
                    'cached': True,
                })
        except Exception:
            pass

    all_odds = {}
    errors = []
    sources_used = []

    # ── Source 1: The Odds API (free tier, most reliable, covers soccer) ──
    if api_key:
        toa_odds, toa_errors = _fetch_the_odds_api(sports, api_key)
        if toa_odds:
            all_odds.update(toa_odds)
            sources_used.append('The Odds API')
        errors.extend(toa_errors)

    # ── Source 2: Bovada (fills gaps for sports it covers) ──
    bovada_sports = [s for s in sports if s in BOVADA_FUTURES]
    if bovada_sports:
        bovada_odds, bovada_errors = _fetch_bovada(bovada_sports)
        for k, v in bovada_odds.items():
            if k not in all_odds:   # don't overwrite The Odds API data
                all_odds[k] = v
        if bovada_odds:
            sources_used.append('Bovada')
        errors.extend(bovada_errors)

    # ── Source 3: ESPN (fills remaining gaps) ──
    espn_sports = [s for s in sports if s in SPORT_ENDPOINTS]
    if espn_sports:
        espn_odds, espn_errors = _fetch_espn(espn_sports)
        for k, v in espn_odds.items():
            if k not in all_odds:   # don't overwrite higher-priority sources
                all_odds[k] = v
        if espn_odds:
            sources_used.append('ESPN')
        errors.extend(espn_errors)

    source = ' + '.join(sources_used) if sources_used else 'none'

    # Persist snapshot for line movement tracking
    if all_odds:
        _append_odds_to_history(all_odds)

    response_data = {
        'odds':   all_odds,
        'count':  len(all_odds),
        'source': source,
        'errors': errors or None,
    }
    _futures_cache_set(cache_key, response_data)

    return jsonify({
        'ok': True,
        **response_data,
        'cached': False,
    })


# ── futures auto-settlement ────────────────────────────────────────────────
# Detects open futures whose championship event has ended (per
# futures_event_dates.json) and lets the dashboard one-click settle them
# with W/L/P. Settlement writes back to the user's xlsx — row moves from
# the "Open Bets" sheet to the "Bet History" sheet.

@app.route('/api/stale-futures', methods=['GET'])
def stale_futures():
    """Return open futures for ?user= whose event has already ended.
    Each item carries eventEndDate + daysPast + championshipName so the UI
    can render a 'needs review' banner with context."""
    user = _get_user(request)
    try:
        from futures_engine import find_stale_futures
        open_bets = read_open_bets(user)
        stale = find_stale_futures(open_bets)
        return jsonify({
            'ok': True,
            'user': user,
            'count': len(stale),
            'futures': stale,
        })
    except ImportError:
        return jsonify({'ok': False, 'error': 'futures_engine.py not found'}), 500
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/settle-bet', methods=['POST'])
def settle_bet_endpoint():
    """Move a row from the 'Open Bets' sheet to the 'Bet History' sheet for
    the current user, marking it Won/Lost/Push. Body:
        {betId: str, result: 'W'|'L'|'P', winningTeam?: str}
    Returns the settled tx_id + computed win/loss amount.
    Invalidates the user's xlsx cache so the next read picks up the move."""
    user = _get_user(request)
    try:
        from futures_engine import settle_bet as engine_settle
        data = request.get_json(force=True) or {}
        bet_id = str(data.get('betId') or '').strip()
        result = (data.get('result') or '').strip().upper()
        winning_team = (data.get('winningTeam') or '').strip() or None

        if not bet_id:
            return jsonify({'ok': False, 'error': 'betId is required'}), 400
        if result not in ('W', 'L', 'P'):
            return jsonify({'ok': False, 'error': "result must be 'W', 'L', or 'P'"}), 400

        path = tracker_path_for(user)
        info = engine_settle(path, bet_id, result, winning_team=winning_team)
        _invalidate_xlsx_cache(user)
        return jsonify({'ok': True, 'user': user, **info})
    except LookupError as e:
        return jsonify({'ok': False, 'error': str(e), 'code': 'BET_NOT_FOUND'}), 404
    except PermissionError as e:
        return jsonify({'ok': False, 'error': 'Workbook locked — close Excel and retry.', 'code': 'XLSX_LOCKED'}), 503
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/closing-lines', methods=['POST'])
def capture_closing_lines():
    """
    Accept a list of bet IDs + their current odds from the frontend.
    Store as closing line snapshots for CLV calculation.
    Body: { "lines": [ { "betId": "...", "closingOdds": -115 }, ... ] }
    """
    try:
        data = request.get_json(force=True)
        lines = data.get('lines', [])
        if not lines:
            return jsonify({'ok': True, 'saved': 0})

        # Load existing closing lines
        cl_file = os.path.join(SCRIPT_DIR, 'closing_lines.json')
        existing = {}
        if os.path.exists(cl_file):
            try:
                with open(cl_file) as f:
                    existing = json.load(f)
            except:
                pass

        saved = 0
        ts = datetime.utcnow().isoformat() + 'Z'
        for item in lines:
            bet_id = item.get('betId', '')
            closing = item.get('closingOdds')
            if bet_id and closing is not None:
                existing[bet_id] = {
                    'closingOdds': closing,
                    'ts': ts,
                }
                saved += 1

        with open(cl_file, 'w') as f:
            json.dump(existing, f, indent=1)

        return jsonify({'ok': True, 'saved': saved})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/closing-lines', methods=['GET'])
def get_closing_lines():
    """Return stored closing lines for CLV display."""
    cl_file = os.path.join(SCRIPT_DIR, 'closing_lines.json')
    if os.path.exists(cl_file):
        try:
            with open(cl_file) as f:
                data = json.load(f)
            return jsonify({'ok': True, 'lines': data, 'count': len(data)})
        except:
            pass
    return jsonify({'ok': True, 'lines': {}, 'count': 0})


@app.route('/api/odds-history', methods=['GET'])
def get_odds_history():
    """
    Return stored odds history for line movement display.
    Query params:
        teams  comma-separated team names (optional, returns all if omitted)
        days   number of days of history (default: 30)
    """
    history = _load_odds_history()
    teams_param = request.args.get('teams', '').strip()
    try:
        days = int(request.args.get('days', 30))
    except (ValueError, TypeError):
        days = 30
    cutoff = (datetime.utcnow() - timedelta(days=days)).isoformat()

    if teams_param:
        team_keys = [t.strip().lower() for t in teams_param.split(',') if t.strip()]
        filtered = {}
        for tk in team_keys:
            for hk in history:
                if tk in hk or hk in tk:
                    entries = [e for e in history[hk] if e.get('ts', '') >= cutoff]
                    if entries:
                        filtered[hk] = entries
        history = filtered
    else:
        history = {
            k: [e for e in v if e.get('ts', '') >= cutoff]
            for k, v in history.items()
            if any(e.get('ts', '') >= cutoff for e in v)
        }

    return jsonify({'ok': True, 'history': history, 'count': len(history)})


# ── upcoming games ─────────────────────────────────────────────────────────────
UPCOMING_CACHE_FILE = os.path.join(SCRIPT_DIR, 'upcoming_games_cache.json')

@app.route('/api/upcoming-games', methods=['GET'])
def upcoming_games():
    """
    Serve pre-cached upcoming games (written by fetch_upcoming_games.py at 2:30 AM).
    Falls back to a live ESPN fetch if the cache file is missing or stale (> 4 hours old).
    """
    import urllib.request as urlreq

    # Try to serve from disk cache first
    if os.path.exists(UPCOMING_CACHE_FILE):
        try:
            with open(UPCOMING_CACHE_FILE) as f:
                data = json.load(f)
            fetched_at_ms = data.get('fetchedAt', 0)
            age_hours = (time.time() * 1000 - fetched_at_ms) / (3600 * 1000)
            if age_hours < 4:
                return jsonify({'ok': True, 'source': 'cache', **data})
        except Exception as e:
            print(f'[upcoming] cache read error: {e}')

    # Live fallback — call fetch_upcoming_games.py as a subprocess
    try:
        fetch_script = os.path.join(SCRIPT_DIR, 'fetch_upcoming_games.py')
        subprocess.run([sys.executable, fetch_script], timeout=30, check=True)
        with open(UPCOMING_CACHE_FILE) as f:
            data = json.load(f)
        return jsonify({'ok': True, 'source': 'live', **data})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'games': []}), 500


# ── CLV (Closing Line Value) endpoints ────────────────────────────────────────

GAME_ODDS_SNAPSHOTS_FILE = os.path.join(SCRIPT_DIR, 'game_odds_snapshots.json')
GAME_ODDS_CONFIG_FILE = os.path.join(SCRIPT_DIR, 'game_odds_config.json')


def _load_game_odds_config():
    if os.path.exists(GAME_ODDS_CONFIG_FILE):
        try:
            with open(GAME_ODDS_CONFIG_FILE) as f:
                return json.load(f)
        except:
            pass
    return {}


@app.route('/api/clv/stats', methods=['GET'])
def clv_stats():
    """Calculate and return aggregate CLV statistics for all settled bets.
    Query params:
        sport   — filter by sport (optional)
        market  — filter by market: h2h, spreads, totals (optional)
    """
    try:
        from clv_calculator import calculate_clv_for_bets, aggregate_clv_stats, load_closing_lines_data

        user = _get_user(request)
        bets = read_settled_bets(user)
        closing = load_closing_lines_data()

        if not closing:
            return jsonify({'ok': True, 'stats': None, 'message': 'No closing line data yet'})

        # Optional filters
        sport_filter = request.args.get('sport', '').strip()
        market_filter = request.args.get('market', '').strip()

        # Calculate CLV for all bets
        clv_bets = calculate_clv_for_bets(bets, closing)

        # Apply filters
        if sport_filter:
            clv_bets = [b for b in clv_bets if (b.get('sport', '') or '').upper() == sport_filter.upper()]
        if market_filter:
            clv_bets = [b for b in clv_bets if b.get('clv_market', '') == market_filter]

        stats = aggregate_clv_stats(clv_bets)
        return jsonify({'ok': True, 'stats': stats})
    except ImportError:
        return jsonify({'ok': False, 'error': 'clv_calculator.py not found'}), 500
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/clv/bets', methods=['GET'])
def clv_bets():
    """Return all settled bets enriched with CLV data.
    Each bet gets: clv_pct, closing_odds, opening_odds, clv_matched, clv_source.
    """
    try:
        from clv_calculator import calculate_clv_for_bets, load_closing_lines_data

        user = _get_user(request)
        bets = read_settled_bets(user)
        closing = load_closing_lines_data()

        if not closing:
            return jsonify({'ok': True, 'bets': bets, 'message': 'No closing line data yet'})

        clv_bets = calculate_clv_for_bets(bets, closing)
        return jsonify({'ok': True, 'bets': clv_bets, 'count': len(clv_bets)})
    except ImportError:
        return jsonify({'ok': False, 'error': 'clv_calculator.py not found'}), 500
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/clv/game-odds-status', methods=['GET'])
def game_odds_status():
    """Return status of the game odds snapshot system.
    Shows: active events, total snapshots, last capture time, credits remaining.
    """
    try:
        snapshots = {}
        if os.path.exists(GAME_ODDS_SNAPSHOTS_FILE):
            with open(GAME_ODDS_SNAPSHOTS_FILE) as f:
                snapshots = json.load(f)

        cl_file = os.path.join(SCRIPT_DIR, 'closing_lines.json')
        closing_count = 0
        if os.path.exists(cl_file):
            with open(cl_file) as f:
                closing_count = len(json.load(f))

        # Find most recent snapshot timestamp
        last_capture = None
        for event in snapshots.values():
            for snap in event.get('snapshots', []):
                ts = snap.get('ts', '')
                if ts and (last_capture is None or ts > last_capture):
                    last_capture = ts

        total_events = len(snapshots)
        total_snaps = sum(len(e.get('snapshots', [])) for e in snapshots.values())

        # Sports breakdown
        sports = {}
        for event in snapshots.values():
            sport = event.get('sport', 'Other')
            sports[sport] = sports.get(sport, 0) + 1

        config = _load_game_odds_config()

        return jsonify({
            'ok': True,
            'active_events': total_events,
            'total_snapshots': total_snaps,
            'closing_lines_count': closing_count,
            'last_capture': last_capture,
            'sports_breakdown': sports,
            'api_key_configured': bool(config.get('api_key') or os.environ.get('THE_ODDS_API_KEY')),
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/clv/config', methods=['GET', 'POST'])
def clv_config():
    """GET: Return current CLV config (api key masked).
    POST: Update CLV config (api key, poll interval, etc.)
    """
    if request.method == 'GET':
        config = _load_game_odds_config()
        # Mask the API key for security
        key = config.get('api_key', '')
        if key:
            config['api_key_masked'] = key[:6] + '...' + key[-4:] if len(key) > 10 else '***'
        else:
            config['api_key_masked'] = ''
        config.pop('api_key', None)
        return jsonify({'ok': True, 'config': config})

    # POST — save config
    try:
        data = request.get_json(force=True)
        config = _load_game_odds_config()

        if 'api_key' in data:
            config['api_key'] = data['api_key']
        if 'poll_interval_hours' in data:
            config['poll_interval_hours'] = data['poll_interval_hours']
        if 'active_sports' in data:
            config['active_sports'] = data['active_sports']

        with open(GAME_ODDS_CONFIG_FILE, 'w') as f:
            json.dump(config, f, indent=2)

        return jsonify({'ok': True, 'message': 'Config saved'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/clv/refresh', methods=['POST'])
def clv_refresh():
    """Trigger a manual game odds snapshot capture.
    Body (optional): { "closing_only": true, "sport": "NBA" }
    Returns a 'status' block with the refresh_game_odds.py exit code
    (exit 5 = daily budget cap hit — see SCRAPER_EXIT_LABELS).
    """
    try:
        data = request.get_json(force=True) if request.data else {}
        closing_only = data.get('closing_only', False)
        sport = data.get('sport')

        cmd = [sys.executable, os.path.join(SCRIPT_DIR, 'refresh_game_odds.py')]
        if closing_only:
            cmd.append('--closing-only')
        if sport:
            cmd.extend(['--sport', sport])

        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        status = _exit_info(result.returncode)

        return jsonify({
            'ok': result.returncode == 0,
            'status': status,
            'stdout': result.stdout[-2000:] if result.stdout else '',
            'stderr': result.stderr[-500:] if result.stderr else '',
        })
    except subprocess.TimeoutExpired:
        return jsonify({'ok': False, 'error': 'Snapshot timed out (120s)',
                        'status': {'code': 2, 'slug': 'scrape', 'label': 'Timed out after 120s', 'ok': False}}), 504
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── Odds API budget ───────────────────────────────────────────────────────────
# Exposes the contents of odds_api_state.json so the dashboard can show daily
# credit usage before the user clicks "Refresh CLV". Purely read-only.
@app.route('/api/budget', methods=['GET'])
def budget_status():
    """Return today's Odds API usage + remaining cap headroom.

    Response shape:
      {
        ok: true,
        today: { credits_used: N, last_remaining: M, date: 'YYYY-MM-DD' },
        daily_cap: 1000,
        remaining_in_cap: <cap - used, clamped at 0>,
        pct_used: <0-100>,
        total_last_remaining: <plan-level remaining reported by API>,
        last_call_ts: <ISO timestamp or null>,
        state_file_present: bool
      }

    Never 500s on a missing / corrupt state file — empty state is a valid
    "no calls today" answer, not an error.
    """
    from datetime import date
    today = date.today().isoformat()
    daily_cap = 1000  # mirrors DEFAULT_DAILY_BUDGET in refresh_game_odds.py

    state = {'daily': {}, 'total_last_remaining': None, 'last_call_ts': None}
    state_present = os.path.exists(ODDS_STATE_FILE)
    if state_present:
        try:
            with open(ODDS_STATE_FILE) as f:
                loaded = json.load(f)
            if isinstance(loaded, dict):
                state.update(loaded)
        except (json.JSONDecodeError, OSError):
            # Treat corrupt state as empty — don't fail the dashboard render.
            pass

    today_entry = state.get('daily', {}).get(today, {}) or {}
    try:
        used = int(today_entry.get('credits_used', 0) or 0)
    except (TypeError, ValueError):
        used = 0
    try:
        last_remaining = int(today_entry.get('last_remaining', 0) or 0)
    except (TypeError, ValueError):
        last_remaining = 0

    remaining = max(0, daily_cap - used)
    pct = min(100, round((used / daily_cap) * 100)) if daily_cap > 0 else 0

    return jsonify({
        'ok': True,
        'today': {
            'date': today,
            'credits_used': used,
            'last_remaining': last_remaining,
        },
        'daily_cap': daily_cap,
        'remaining_in_cap': remaining,
        'pct_used': pct,
        'total_last_remaining': state.get('total_last_remaining'),
        'last_call_ts': state.get('last_call_ts'),
        'state_file_present': state_present,
    })


def _preflight():
    """Sanity-check the environment before Flask binds the socket.
    Checks every per-user xlsx file. Returns (ok, messages).
    Fails only when the DEFAULT_USER's workbook is missing."""
    msgs = []
    ok = True
    for user in ALLOWED_USERS:
        path = tracker_path_for(user)
        if not os.path.exists(path):
            severity = "FATAL" if user == DEFAULT_USER else "WARN"
            msgs.append(f"[{severity}] Missing {path} — create an empty workbook with 'Bet History' and 'Open Bets' sheets.")
            if user == DEFAULT_USER:
                ok = False
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if 'Bet History' not in wb.sheetnames or 'Open Bets' not in wb.sheetnames:
                msgs.append(f"[WARN] {user}'s workbook is missing one of ['Bet History','Open Bets']. Found: {wb.sheetnames}")
            wb.close()
        except PermissionError:
            msgs.append(f"[WARN] Excel lock detected on {path}. Close it in Excel before reads.")
        except Exception as e:
            msgs.append(f"[WARN] Could not open {user}'s workbook for preflight: {e.__class__.__name__}: {e}")
    for script_path, label in ((LOCKS_SCRIPT, 'locks25'), (BOVADA_SCRIPT, 'bovada')):
        if not os.path.exists(script_path):
            msgs.append(f"[WARN] Scraper {label} script not found at {script_path} — refresh endpoint will no-op.")
    return ok, msgs


def _pick_port(default_port):
    """Pick a listen port: env override wins; else default; else first free above default."""
    env = os.environ.get('BETTING_TRACKER_PORT', '').strip()
    if env.isdigit():
        return int(env)
    import socket
    port = default_port
    for attempt in range(5):
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        try:
            s.bind(('127.0.0.1', port))
            s.close()
            return port
        except OSError:
            s.close()
            port += 1
    return default_port  # hand back to Flask; it'll error cleanly


if __name__ == '__main__':
    ok, preflight_msgs = _preflight()
    port = _pick_port(5001)
    print("=" * 55)
    print("  Betting Tracker — Data Bridge Server")
    print(f"  Users: {', '.join(ALLOWED_USERS)} (default: {DEFAULT_USER})")
    print(f"  Listening on http://localhost:{port}")
    if port != 5001:
        print(f"  (default 5001 was busy — override via BETTING_TRACKER_PORT=5001 once free)")
    print("  Keep this running — dashboard auto-connects")
    for m in preflight_msgs:
        print("  " + m)
    print("=" * 55)
    if not ok:
        sys.exit(1)
    app.run(host='127.0.0.1', port=port, debug=False)
